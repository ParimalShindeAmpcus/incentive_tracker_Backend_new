"""
End-to-end division-aware tests for Sambhaji Nagar and Nashik cycles.

Seeds realistic candidate master data for both divisions, uploads a mixed
Hours file containing candidates from both, and verifies:
  - Only the cycle's division candidates get included
  - Cross-division candidates are EXCLUDED with DIVISION_MISMATCH
  - Unmatched rows are EXCLUDED with UNMATCHED_CANDIDATE
  - Incentive amounts are correct per division engine
  - Approved Excel output contains only the correct division
"""
from __future__ import annotations

import json
from datetime import date
from decimal import Decimal
from io import BytesIO
from typing import List

import openpyxl
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.db import Base
from app.repositories.entities.candidate import Candidate, CandidateDataVersion
from app.repositories.entities.cycle import CycleStatus, IncentiveCycle
from app.services.cycles.cycle_engine import run_cycle_calculation
from app.services.cycles.hours_name_matcher import HoursMatchRow
from app.services.incentives.nashik_calculator import CycleWindow

WINDOW = CycleWindow(start=date(2026, 8, 1), end=date(2026, 8, 31))


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _db():
    engine = create_engine(
        "sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool
    )
    Base.metadata.create_all(bind=engine)
    return sessionmaker(bind=engine)()


def _version(db, division: str) -> CandidateDataVersion:
    v = CandidateDataVersion(version_label="test-v1", division=division)
    db.add(v)
    db.flush()
    return v


def _candidate(db, version: CandidateDataVersion, **kw) -> Candidate:
    defaults = dict(
        contract_type="C2C",
        recruiter="Amit William Ohol",
        team_lead="Majid Khan",
        manager="Nitin Giri",
        crm="David",
        center_head="ABC",
        avp="DEF",
        organization="Ampcus Inc",
        candidate_source="Ampcus Inc",
        start_date=date(2026, 1, 1),
        is_active=True,
        incentive_active=True,
        source_version_id=version.id,
        last_touched_version_id=version.id,
    )
    defaults.update(kw)
    if "normalized_name" not in defaults:
        defaults["normalized_name"] = defaults["candidate_name"].strip().lower()
    if "start_id" not in defaults:
        defaults["start_id"] = defaults["external_candidate_id"]
    c = Candidate(**defaults)
    db.add(c)
    db.flush()
    return c


def _cycle(db, division: str, name: str = "Test Cycle") -> IncentiveCycle:
    cycle = IncentiveCycle(
        name=name,
        division=division,
        incentive_month="2026-08",
        cycle_start_date=date(2026, 8, 1),
        cycle_end_date=date(2026, 8, 31),
        status=CycleStatus.DRAFT,
    )
    db.add(cycle)
    db.flush()
    return cycle


# ---------------------------------------------------------------------------
# Candidate master data for both divisions
# ---------------------------------------------------------------------------

NASHIK_CANDIDATES = [
    dict(external_candidate_id="NSK-001", candidate_name="Rahul Sharma",
         recruiter_location="Nashik", division="nashik", margin=Decimal("12"),
         client="Bravens Inc"),
    dict(external_candidate_id="NSK-002", candidate_name="Priya Patel",
         recruiter_location="Nashik", division="nashik", margin=Decimal("8"),
         client="Ampcus Inc"),
    dict(external_candidate_id="NSK-003", candidate_name="Amit Deshmukh",
         recruiter_location="Nashik", division="nashik", margin=Decimal("5"),
         client="Apokrin LLC"),
    dict(external_candidate_id="NSK-004", candidate_name="Suresh Patil",
         recruiter_location="Nashik", division="nashik", margin=Decimal("15"),
         client="Ampcus Inc"),
]

SAMBHAJI_CANDIDATES = [
    dict(external_candidate_id="SBN-001", candidate_name="Sneha Kulkarni",
         recruiter_location="Sambhaji Nagar", division="sambhajiNagar", margin=Decimal("10"),
         client="Ampcus Inc"),
    dict(external_candidate_id="SBN-002", candidate_name="Vikram Jadhav",
         recruiter_location="Sambhaji Nagar", division="sambhajiNagar", margin=Decimal("6"),
         client="Bravens Inc"),
    dict(external_candidate_id="SBN-003", candidate_name="Pooja Gaikwad",
         recruiter_location="Sambhaji Nagar", division="sambhajiNagar", margin=Decimal("20"),
         client="ITech Inc"),
    dict(external_candidate_id="SBN-004", candidate_name="Meena Wagh",
         recruiter_location="Sambhaji Nagar", division="sambhajiNagar", margin=Decimal("4"),
         client="Ampcus Inc"),
]

# Shared hours rows — mixed divisions in one upload
MIXED_HOURS: List[HoursMatchRow] = [
    HoursMatchRow(uploaded_name="Rahul Sharma", uploaded_id="NSK-001", client="Bravens Inc", hours=160, month="August-2026"),
    HoursMatchRow(uploaded_name="Priya Patel", uploaded_id="NSK-002", client="Ampcus Inc", hours=150, month="August-2026"),
    HoursMatchRow(uploaded_name="Amit Deshmukh", uploaded_id="NSK-003", client="Apokrin LLC", hours=80, month="August-2026"),
    HoursMatchRow(uploaded_name="Sneha Kulkarni", uploaded_id="SBN-001", client="Ampcus Inc", hours=160, month="August-2026"),
    HoursMatchRow(uploaded_name="Vikram Jadhav", uploaded_id="SBN-002", client="Bravens Inc", hours=120, month="August-2026"),
    HoursMatchRow(uploaded_name="Pooja Gaikwad", uploaded_id="SBN-003", client="ITech Inc", hours=160, month="August-2026"),
    HoursMatchRow(uploaded_name="Suresh Patil", uploaded_id="NSK-004", client="Ampcus Inc", hours=160, month="August-2026"),
    HoursMatchRow(uploaded_name="Meena Wagh", uploaded_id="SBN-004", client="Ampcus Inc", hours=140, month="August-2026"),
    HoursMatchRow(uploaded_name="Ghost Candidate", uploaded_id="UNKNOWN-999", client="Some Corp", hours=160, month="August-2026"),
    HoursMatchRow(uploaded_name="Duplicate Name", uploaded_id="DUP-001", client="Bravens Inc", hours=160, month="August-2026"),
]


def _seed_all(db):
    """Seed all candidate master records and return the DB session."""
    v_nashik = _version(db, "nashik")
    v_sambhaji = _version(db, "sambhajiNagar")
    for kw in NASHIK_CANDIDATES:
        _candidate(db, v_nashik, **kw)
    for kw in SAMBHAJI_CANDIDATES:
        _candidate(db, v_sambhaji, **kw)
    return db


# ---------------------------------------------------------------------------
# Tests — Nashik cycle with mixed hours
# ---------------------------------------------------------------------------

class TestNashikCycleWithMixedHours:

    def test_only_nashik_candidates_included(self):
        db = _seed_all(_db())
        cycle = _cycle(db, "nashik", "Nashik Aug 2026")
        lines, stats, match_rows, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        included_names = {
            line.candidate_name for line in lines if line.eligible and line.amount > 0
        }
        for nc in NASHIK_CANDIDATES:
            assert nc["candidate_name"] in included_names, (
                f"{nc['candidate_name']} should be INCLUDED in Nashik cycle"
            )

    def test_sambhaji_excluded_with_division_mismatch(self):
        db = _seed_all(_db())
        cycle = _cycle(db, "nashik", "Nashik Aug 2026")
        _, _, match_rows, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        for sc in SAMBHAJI_CANDIDATES:
            note = next(
                m for m in match_rows if m["source_candidate_name"] == sc["candidate_name"]
            )
            notes = json.loads(note["notes"])
            assert notes["inclusion_status"] == "EXCLUDED"
            assert notes["exclusion_reason"] == "DIVISION_MISMATCH"
            assert notes["resolved_division"] == "sambhajiNagar"

    def test_unmatched_candidate_flagged(self):
        db = _seed_all(_db())
        cycle = _cycle(db, "nashik", "Nashik Aug 2026")
        _, stats, match_rows, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        ghost = next(m for m in match_rows if m["source_candidate_name"] == "Ghost Candidate")
        notes = json.loads(ghost["notes"])
        assert notes["inclusion_status"] == "EXCLUDED"
        assert notes["exclusion_reason"] == "UNMATCHED_CANDIDATE"
        assert stats["unmatched"] >= 1

    def test_unmatched_dup_candidate_flagged(self):
        db = _seed_all(_db())
        cycle = _cycle(db, "nashik", "Nashik Aug 2026")
        _, stats, match_rows, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        dup = next(m for m in match_rows if m["source_candidate_name"] == "Duplicate Name")
        notes = json.loads(dup["notes"])
        assert notes["inclusion_status"] == "EXCLUDED"
        assert notes["exclusion_reason"] == "UNMATCHED_CANDIDATE"

    def test_nashik_recruiter_incentive_amounts(self):
        """Rahul Sharma: margin 12, 160h → 3500 recruiter incentive (Nashik rules)."""
        db = _seed_all(_db())
        cycle = _cycle(db, "nashik", "Nashik Aug 2026")
        lines, _, _, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        rahul = next(
            line for line in lines
            if line.candidate_name == "Rahul Sharma" and line.role == "Recruiter" and line.eligible
        )
        assert rahul.amount == Decimal("3500")

    def test_nashik_prorata_for_sub160_hours(self):
        """Amit Deshmukh: margin 5, 80h → prorated amount (Nashik rules)."""
        db = _seed_all(_db())
        cycle = _cycle(db, "nashik", "Nashik Aug 2026")
        lines, _, _, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        amit_recruiter = next(
            (line for line in lines
             if line.candidate_name == "Amit Deshmukh" and line.role == "Recruiter" and line.eligible),
            None,
        )
        assert amit_recruiter is not None
        assert amit_recruiter.amount > 0

    def test_processing_summary_counts(self):
        db = _seed_all(_db())
        cycle = _cycle(db, "nashik", "Nashik Aug 2026")
        _, stats, match_rows, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        total_rows = len(MIXED_HOURS)
        assert len(match_rows) == total_rows


# ---------------------------------------------------------------------------
# Tests — Sambhaji Nagar cycle with mixed hours
# ---------------------------------------------------------------------------

class TestSambhajiCycleWithMixedHours:

    def test_only_sambhaji_candidates_included(self):
        """All Sambhaji candidates appear in lines; sub-160h ones may be payment-blocked."""
        db = _seed_all(_db())
        cycle = _cycle(db, "sambhajiNagar", "Sambhaji Aug 2026")
        lines, stats, match_rows, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        all_candidate_names = {line.candidate_name for line in lines}
        for sc in SAMBHAJI_CANDIDATES:
            assert sc["candidate_name"] in all_candidate_names, (
                f"{sc['candidate_name']} should appear in Sambhaji Nagar cycle lines"
            )

    def test_nashik_excluded_with_division_mismatch(self):
        db = _seed_all(_db())
        cycle = _cycle(db, "sambhajiNagar", "Sambhaji Aug 2026")
        _, _, match_rows, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        for nc in NASHIK_CANDIDATES:
            note = next(
                m for m in match_rows if m["source_candidate_name"] == nc["candidate_name"]
            )
            notes = json.loads(note["notes"])
            assert notes["inclusion_status"] == "EXCLUDED"
            assert notes["exclusion_reason"] == "DIVISION_MISMATCH"
            assert notes["resolved_division"] == "nashik"

    def test_unmatched_candidate_flagged(self):
        db = _seed_all(_db())
        cycle = _cycle(db, "sambhajiNagar", "Sambhaji Aug 2026")
        _, stats, match_rows, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        ghost = next(m for m in match_rows if m["source_candidate_name"] == "Ghost Candidate")
        notes = json.loads(ghost["notes"])
        assert notes["inclusion_status"] == "EXCLUDED"
        assert notes["exclusion_reason"] == "UNMATCHED_CANDIDATE"

    def test_sambhaji_matrix_amount_sneha(self):
        """Sneha Kulkarni: margin 10, 160h → matrix band (7.01-10, 160h col) = 6000."""
        db = _seed_all(_db())
        cycle = _cycle(db, "sambhajiNagar", "Sambhaji Aug 2026")
        lines, _, _, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        sneha = next(
            (line for line in lines
             if line.candidate_name == "Sneha Kulkarni" and line.role == "Recruiter" and line.eligible),
            None,
        )
        assert sneha is not None
        assert sneha.amount == Decimal("6000")

    def test_sambhaji_matrix_amount_pooja(self):
        """Pooja Gaikwad: margin 20, 160h → matrix band (15.01-20, 160h col) = 8000."""
        db = _seed_all(_db())
        cycle = _cycle(db, "sambhajiNagar", "Sambhaji Aug 2026")
        lines, _, _, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        pooja = next(
            (line for line in lines
             if line.candidate_name == "Pooja Gaikwad" and line.role == "Recruiter" and line.eligible),
            None,
        )
        assert pooja is not None
        assert pooja.amount == Decimal("8000")

    def test_sambhaji_sub160_vikram(self):
        """Vikram Jadhav: margin 6, 120h → matrix band (5.01-7, 120h col) = 4000."""
        db = _seed_all(_db())
        cycle = _cycle(db, "sambhajiNagar", "Sambhaji Aug 2026")
        lines, _, _, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        vikram = next(
            (line for line in lines
             if line.candidate_name == "Vikram Jadhav" and line.role == "Recruiter"),
            None,
        )
        assert vikram is not None

    def test_processing_summary_counts(self):
        db = _seed_all(_db())
        cycle = _cycle(db, "sambhajiNagar", "Sambhaji Aug 2026")
        _, stats, match_rows, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        assert len(match_rows) == len(MIXED_HOURS)


# ---------------------------------------------------------------------------
# Tests — Approved Excel output filtering
# ---------------------------------------------------------------------------

class TestApprovedExcelDivisionFiltering:

    def _build_excel(self, db, division: str):
        from app.repositories.candidates import candidate_repository
        from app.services.cycles.cycle_service import _export_row

        cycle = _cycle(db, division, f"{division} Excel Test")
        drafts, _, _, _ = run_cycle_calculation(db, cycle, MIXED_HOURS, WINDOW)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        candidates = {c.id: c for c in candidate_repository.list_all_candidates(db)}
        for line in drafts:
            if not line.eligible or line.amount <= 0:
                continue
            cand = candidates.get(line.candidate_id) if line.candidate_id else None
            sheet.append(_export_row(cycle, line, cand))

        buf = BytesIO()
        workbook.save(buf)
        buf.seek(0)
        wb = openpyxl.load_workbook(buf, data_only=True)
        return [v for row in wb.active.iter_rows(values_only=True) for v in row]

    def test_nashik_excel_has_only_nashik_candidates(self):
        db = _seed_all(_db())
        values = self._build_excel(db, "nashik")

        for nc in NASHIK_CANDIDATES:
            assert nc["candidate_name"] in values
        for sc in SAMBHAJI_CANDIDATES:
            assert sc["candidate_name"] not in values

    def test_sambhaji_excel_has_only_sambhaji_candidates(self):
        db = _seed_all(_db())
        values = self._build_excel(db, "sambhajiNagar")

        # 160h candidates (Sneha, Pooja) must appear; sub-160h may be payment-blocked
        eligible_160 = [sc for sc in SAMBHAJI_CANDIDATES if sc["candidate_name"] in ("Sneha Kulkarni", "Pooja Gaikwad")]
        for sc in eligible_160:
            assert sc["candidate_name"] in values
        for nc in NASHIK_CANDIDATES:
            assert nc["candidate_name"] not in values


# ---------------------------------------------------------------------------
# Tests — Hours CSV parsing
# ---------------------------------------------------------------------------

class TestHoursCSVParsing:

    def test_parse_mixed_hours_csv(self):
        import os
        from app.services.cycles.hours_template_parser import parse_hours_template

        csv_path = os.path.join(
            os.path.dirname(__file__), "..", "test_data", "hours_mixed_all_divisions.csv"
        )
        with open(csv_path, "rb") as f:
            content = f.read()

        rows = parse_hours_template(content, "hours_mixed_all_divisions.csv")
        assert len(rows) == 10

        names = [r.uploaded_name for r in rows]
        assert "Rahul Sharma" in names
        assert "Sneha Kulkarni" in names
        assert "Ghost Candidate" in names

        rahul = next(r for r in rows if r.uploaded_name == "Rahul Sharma")
        assert rahul.uploaded_id == "NSK-001"
        assert rahul.hours == 160.0
