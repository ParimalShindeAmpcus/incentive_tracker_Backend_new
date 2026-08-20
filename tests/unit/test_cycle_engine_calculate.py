from datetime import date
from decimal import Decimal

from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.db import Base
from app.repositories.entities.candidate import Candidate, CandidateDataVersion
from app.repositories.entities.coordinator import CoordinatorRecord, CoordinatorStatus
from app.repositories.entities.cycle import CycleStatus, IncentiveCycle
from app.services.cycles.cycle_engine import run_cycle_calculation
from app.services.cycles.hours_name_matcher import HoursMatchRow
from app.services.incentives.nashik_calculator import CycleWindow
from app.services.incentives.nashik_rules import normalize_person


def _session():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False}, poolclass=StaticPool)
    Base.metadata.create_all(bind=engine)
    return sessionmaker(bind=engine)()


def _seed_candidate(db, **kwargs):
    version_division = kwargs.pop("division", "nashik")
    recruiter_location = kwargs.pop("recruiter_location", "Nashik")
    organization = kwargs.pop("organization", "Ampcus Inc")
    candidate_name = kwargs.pop("candidate_name", "Aisha Mayes")
    normalized_name = kwargs.pop("normalized_name", candidate_name.lower())
    external_candidate_id = kwargs.pop("external_candidate_id", "12345")
    start_id = kwargs.pop("start_id", external_candidate_id)

    version = CandidateDataVersion(version_label="v1", division=version_division)
    db.add(version)
    db.flush()
    data = dict(
        external_candidate_id=external_candidate_id,
        start_id=start_id,
        candidate_name=candidate_name,
        normalized_name=normalized_name,
        contract_type="C2C",
        margin=Decimal("12"),
        recruiter="Amit William Ohol",
        team_lead="Majid Khan",
        manager="Nitin Giri",
        crm="David",
        center_head="ABC",
        avp="DEF",
        organization=organization,
        candidate_source=organization,
        recruiter_location=recruiter_location,
        start_date=date(2026, 1, 1),
        division=version_division,
        source_version_id=version.id,
        last_touched_version_id=version.id,
        incentive_active=True,
        is_active=True,
    )
    data.update(kwargs)
    cand = Candidate(**data)
    db.add(cand)
    db.flush()
    return cand


def _cycle(db, *, division: str = "nashik"):
    cycle = IncentiveCycle(
        name="Nashik Aug",
        division=division,
        incentive_month="2026-08",
        cycle_start_date=date(2026, 8, 1),
        cycle_end_date=date(2026, 8, 31),
        status=CycleStatus.DRAFT,
    )
    db.add(cycle)
    db.flush()
    return cycle


WINDOW = CycleWindow(start=date(2026, 8, 1), end=date(2026, 8, 31))


def test_calculate_creates_lines_for_name_and_id_match():
    db = _session()
    _seed_candidate(db)
    cycle = _cycle(db)
    rows = [HoursMatchRow(uploaded_name="Aisha Mayes", uploaded_id="12345", hours=160)]
    lines, stats, _, _ = run_cycle_calculation(db, cycle, rows, WINDOW)
    recruiter = next(line for line in lines if line.role == "Recruiter" and line.eligible)
    assert recruiter.amount == Decimal("3500")
    assert stats["matched_name_and_id"] == 1
    assert any(line.role == "Team Lead" and line.amount == Decimal("250") for line in lines)
    assert any(line.role == "Manager" and line.eligible for line in lines)


def test_name_id_mismatch_does_not_pay():
    db = _session()
    _seed_candidate(db)
    cycle = _cycle(db)
    rows = [HoursMatchRow(uploaded_name="Aisha Mayes", uploaded_id="99999", hours=160)]
    lines, stats, _, _ = run_cycle_calculation(db, cycle, rows, WINDOW)
    assert stats["name_id_mismatch"] == 1
    assert all((not line.eligible) or line.amount == 0 for line in lines)


def test_eighty_hours_prorata():
    db = _session()
    _seed_candidate(db, margin=Decimal("8"))
    cycle = _cycle(db)
    rows = [HoursMatchRow(uploaded_name="Aisha Mayes", uploaded_id="12345", hours=80)]
    lines, _, _, _ = run_cycle_calculation(db, cycle, rows, WINDOW)
    recruiter = next(line for line in lines if line.role == "Recruiter" and line.eligible)
    assert recruiter.amount == Decimal("1000")
    tl = next(line for line in lines if line.role == "Team Lead")
    assert tl.amount == Decimal("125")


def test_project_end_before_160():
    db = _session()
    _seed_candidate(db, end_date=date(2026, 8, 10), margin=Decimal("8"))
    cycle = _cycle(db)
    rows = [HoursMatchRow(uploaded_name="Aisha Mayes", uploaded_id="12345", hours=80)]
    lines, _, _, _ = run_cycle_calculation(db, cycle, rows, WINDOW)
    recruiter = next(line for line in lines if line.role == "Recruiter")
    assert recruiter.amount == Decimal("2000")
    assert recruiter.incentive_type == "SPECIAL"
    assert all(line.role != "Team Lead" or line.amount == 0 for line in lines)


def test_missing_hierarchy_does_not_invent_roles():
    db = _session()
    _seed_candidate(db, team_lead=None, manager=None, crm=None, center_head=None, avp=None, margin=Decimal("8"))
    cycle = _cycle(db)
    rows = [HoursMatchRow(uploaded_name="Aisha Mayes", uploaded_id="12345", hours=160)]
    lines, _, _, _ = run_cycle_calculation(db, cycle, rows, WINDOW)
    roles = {line.role for line in lines if line.eligible and line.amount > 0}
    assert roles == {"Recruiter"}


def test_id_fallback_still_calculates():
    db = _session()
    _seed_candidate(db)
    cycle = _cycle(db)
    rows = [HoursMatchRow(uploaded_name="Aisha Mays", uploaded_id="12345", hours=160)]
    lines, stats, _, _ = run_cycle_calculation(db, cycle, rows, WINDOW)
    assert stats["matched_id_fallback"] == 1
    recruiter = next(line for line in lines if line.role == "Recruiter" and line.eligible)
    assert recruiter.amount == Decimal("3500")


def test_unmatched_does_not_pay():
    db = _session()
    _seed_candidate(db)
    cycle = _cycle(db)
    rows = [HoursMatchRow(uploaded_name="Unknown Candidate", uploaded_id="UNKNOWN123", hours=160)]
    lines, stats, _, _ = run_cycle_calculation(db, cycle, rows, WINDOW)
    assert stats["unmatched"] == 1
    assert all((not line.eligible) or line.amount == 0 for line in lines)


def test_explanation_includes_master_start_contract_margin():
    db = _session()
    _seed_candidate(db)
    cycle = _cycle(db)
    rows = [HoursMatchRow(uploaded_name="Aisha Mayes", uploaded_id="12345", hours=160)]
    lines, _, _, _ = run_cycle_calculation(db, cycle, rows, WINDOW)
    recruiter = next(line for line in lines if line.role == "Recruiter" and line.eligible)
    import json

    payload = json.loads(recruiter.explanation[0])
    assert payload["start_date"] == "2026-01-01"
    assert payload["contract_type"] == "C2C"
    assert payload["margin_per_hour"] == 12.0
    assert payload["candidate_id"] == "12345"


def test_division_mismatch_excludes_nashik_from_sambhaji():
    db = _session()
    _seed_candidate(
        db,
        candidate_name="Cand Sambhaji",
        normalized_name="cand sambhaji",
        external_candidate_id="S1",
        start_id="S1",
        recruiter_location="Sambhaji Nagar",
        organization="Ampcus Inc",
        division="sambhajiNagar",
        margin=Decimal("12"),
    )
    _seed_candidate(
        db,
        candidate_name="Cand Nashik",
        normalized_name="cand nashik",
        external_candidate_id="N1",
        start_id="N1",
        recruiter_location="Nashik",
        organization="Ampcus Inc",
        division="nashik",
        margin=Decimal("12"),
    )

    cycle = _cycle(db, division="sambhajiNagar")
    rows = [
        HoursMatchRow(uploaded_name="Cand Sambhaji", uploaded_id="S1", hours=160),
        HoursMatchRow(uploaded_name="Cand Nashik", uploaded_id="N1", hours=160),
    ]

    _, _, match_rows, _ = run_cycle_calculation(db, cycle, rows, WINDOW)

    import json

    mismatch_note = next(m for m in match_rows if m["source_candidate_name"] == "Cand Nashik")
    notes = json.loads(mismatch_note["notes"])
    assert notes["inclusion_status"] == "EXCLUDED"
    assert notes["exclusion_reason"] == "DIVISION_MISMATCH"


def test_division_mismatch_excludes_sambhaji_from_nashik():
    db = _session()
    _seed_candidate(
        db,
        candidate_name="Cand Sambhaji",
        normalized_name="cand sambhaji",
        external_candidate_id="S1",
        start_id="S1",
        recruiter_location="Sambhaji Nagar",
        organization="Ampcus Inc",
        division="sambhajiNagar",
        margin=Decimal("12"),
    )
    _seed_candidate(
        db,
        candidate_name="Cand Nashik",
        normalized_name="cand nashik",
        external_candidate_id="N1",
        start_id="N1",
        recruiter_location="Nashik",
        organization="Ampcus Inc",
        division="nashik",
        margin=Decimal("12"),
    )

    cycle = _cycle(db, division="nashik")
    rows = [
        HoursMatchRow(uploaded_name="Cand Sambhaji", uploaded_id="S1", hours=160),
        HoursMatchRow(uploaded_name="Cand Nashik", uploaded_id="N1", hours=160),
    ]

    _, _, match_rows, _ = run_cycle_calculation(db, cycle, rows, WINDOW)

    import json

    mismatch_note = next(m for m in match_rows if m["source_candidate_name"] == "Cand Sambhaji")
    notes = json.loads(mismatch_note["notes"])
    assert notes["inclusion_status"] == "EXCLUDED"
    assert notes["exclusion_reason"] == "DIVISION_MISMATCH"


def test_unmatched_row_is_reported_as_unmatched_candidate_reason():
    db = _session()
    _seed_candidate(db)
    cycle = _cycle(db, division="nashik")
    rows = [HoursMatchRow(uploaded_name="Unknown Candidate", uploaded_id="UNKNOWN123", hours=160)]

    _, stats, match_rows, _ = run_cycle_calculation(db, cycle, rows, WINDOW)
    assert stats["unmatched"] == 1

    import json

    note = match_rows[0]
    notes = json.loads(note["notes"])
    assert notes["inclusion_status"] == "EXCLUDED"
    assert notes["exclusion_reason"] == "UNMATCHED_CANDIDATE"


def test_approved_excel_excludes_division_mismatch_candidates():
    from io import BytesIO

    import openpyxl

    from app.repositories.cycles import cycle_repository
    from app.repositories.incentives import incentive_repository
    from app.services.cycles.cycle_service import export_cycle

    db = _session()
    _seed_candidate(
        db,
        candidate_name="Cand Sambhaji",
        normalized_name="cand sambhaji",
        external_candidate_id="S1",
        start_id="S1",
        recruiter_location="Sambhaji Nagar",
        organization="Ampcus Inc",
        division="sambhajiNagar",
        margin=Decimal("12"),
    )
    _seed_candidate(
        db,
        candidate_name="Cand Nashik",
        normalized_name="cand nashik",
        external_candidate_id="N1",
        start_id="N1",
        recruiter_location="Nashik",
        organization="Ampcus Inc",
        division="nashik",
        margin=Decimal("12"),
    )

    cycle = _cycle(db, division="sambhajiNagar")
    rows = [
        HoursMatchRow(uploaded_name="Cand Sambhaji", uploaded_id="S1", hours=160),
        HoursMatchRow(uploaded_name="Cand Nashik", uploaded_id="N1", hours=160),
    ]

    drafts, _, match_rows, _ = run_cycle_calculation(db, cycle, rows, WINDOW)

    cycle_repository.replace_matches(db, cycle.id, match_rows)
    incentive_repository.replace_cycle_lines(
        db,
        cycle.id,
        [
            {
                "candidate_id": line.candidate_id,
                "candidate_name": line.candidate_name,
                "role": line.role,
                "person": line.person,
                "incentive_type": line.incentive_type,
                "rule_applied": line.rule_applied,
                "eligible": line.eligible,
                "base_incentive": line.base_incentive,
                "pro_rata_factor": line.pro_rata_factor,
                "amount": line.amount,
                "hours": line.hours,
                "margin": line.margin,
                "reason": line.reason,
                "explanation_json": line.explanation_json(),
                "payment_status": "UNPAID",
            }
            for line in drafts
        ],
    )

    # Recreate the relevant export logic (export_cycle streams a workbook via
    # StreamingResponse, which is awkward to consume in a sync unit test).
    from app.repositories.candidates import candidate_repository
    from app.services.cycles.cycle_service import _export_row

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    candidates = {cand.id: cand for cand in candidate_repository.list_all_candidates(db)}
    # Headers are not needed for assertion; we only care about row values.
    for line in drafts:
        if not line.eligible or line.amount <= 0:
            continue
        cand = candidates.get(line.candidate_id) if line.candidate_id else None
        sheet.append(_export_row(cycle, line, cand))

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    wb = openpyxl.load_workbook(buffer, data_only=True)
    sheet = wb.active
    values = [v for row in sheet.iter_rows(values_only=True) for v in row]

    assert "Cand Sambhaji" in values
    assert "Cand Nashik" not in values


def test_hours_all_divisions_only_selected_cycle_included():
    """
    Hours upload may contain candidates from multiple divisions.
    For a Nashik cycle, only candidates that resolve to Nashik must be included.
    """
    db = _session()

    _seed_candidate(
        db,
        candidate_name="Cand Nashik",
        normalized_name="cand nashik",
        external_candidate_id="N1",
        start_id="N1",
        recruiter_location="Nashik",
        organization="Ampcus Inc",
        division="nashik",
        margin=Decimal("12"),
    )
    _seed_candidate(
        db,
        candidate_name="Cand Sambhaji",
        normalized_name="cand sambhaji",
        external_candidate_id="S1",
        start_id="S1",
        recruiter_location="Sambhaji Nagar",
        organization="Ampcus Inc",
        division="sambhajiNagar",
        margin=Decimal("12"),
    )
    _seed_candidate(
        db,
        candidate_name="Cand Client",
        normalized_name="cand client",
        external_candidate_id="C1",
        start_id="C1",
        recruiter_location="Nashik",
        organization="Ampcus Inc",
        division="ampcusTechClient",
        margin=Decimal("12"),
    )
    _seed_candidate(
        db,
        candidate_name="Cand Inhouse",
        normalized_name="cand inhouse",
        external_candidate_id="I1",
        start_id="I1",
        recruiter_location="Nashik",
        organization="Ampcus Inc",
        division="ampcusTechInhouse",
        margin=Decimal("12"),
    )
    _seed_candidate(
        db,
        candidate_name="Cand FullTime",
        normalized_name="cand fulltime",
        external_candidate_id="F1",
        start_id="F1",
        recruiter_location="Nashik",
        organization="Ampcus Inc",
        division="full_time",
        contract_type="FULLTIME",
        margin=Decimal("12"),
    )

    cycle = _cycle(db, division="nashik")
    rows = [
        HoursMatchRow(uploaded_name="Cand Nashik", uploaded_id="N1", hours=160),
        HoursMatchRow(uploaded_name="Cand Sambhaji", uploaded_id="S1", hours=160),
        HoursMatchRow(uploaded_name="Cand Client", uploaded_id="C1", hours=160),
        HoursMatchRow(uploaded_name="Cand Inhouse", uploaded_id="I1", hours=160),
        HoursMatchRow(uploaded_name="Cand FullTime", uploaded_id="F1", hours=160),
    ]

    lines, stats, match_rows, _ = run_cycle_calculation(db, cycle, rows, WINDOW)

    assert any(line.role == "Recruiter" and line.eligible and line.candidate_name == "Cand Nashik" for line in lines)
    assert not any(line.role == "Recruiter" and line.eligible and line.candidate_name != "Cand Nashik" for line in lines)

    import json

    for src_name in ["Cand Sambhaji", "Cand Client", "Cand Inhouse", "Cand FullTime"]:
        note = next(m for m in match_rows if m["source_candidate_name"] == src_name)
        notes = json.loads(note["notes"])
        assert notes["exclusion_reason"] == "DIVISION_MISMATCH"
        assert notes["inclusion_status"] == "EXCLUDED"

    assert stats["unmatched"] == 0


def _seed_coordinator(db, full_name: str, status: CoordinatorStatus, email_suffix: str = "a"):
    db.add(
        CoordinatorRecord(
            full_name=full_name,
            normalized_name=normalize_person(full_name),
            email=f"{email_suffix}.{normalize_person(full_name).replace(' ', '.')}@example.com",
            organization="Ampcus Inc",
            role_title="Coordinator",
            employment_status=status,
            incentive_eligible=status == CoordinatorStatus.ACTIVE,
        )
    )
    db.flush()


def test_nashik_recruiter_left_hierarchy_continues_via_coordinator_master():
    db = _session()
    _seed_candidate(
        db,
        recruiter="Rec Left",
        team_lead="TL Active",
        manager="Mgr Active",
        crm="CRM Active",
        center_head=None,
        avp=None,
        margin=Decimal("8"),
    )
    _seed_coordinator(db, "Rec Left", CoordinatorStatus.LEFT, "1")
    _seed_coordinator(db, "TL Active", CoordinatorStatus.ACTIVE, "2")
    _seed_coordinator(db, "Mgr Active", CoordinatorStatus.ACTIVE, "3")
    _seed_coordinator(db, "CRM Active", CoordinatorStatus.ACTIVE, "4")
    cycle = _cycle(db)
    rows = [HoursMatchRow(uploaded_name="Aisha Mayes", uploaded_id="12345", hours=160)]
    lines, _, _, _ = run_cycle_calculation(db, cycle, rows, WINDOW)

    rec = next(line for line in lines if line.role == "Recruiter")
    assert rec.eligible is False
    assert rec.amount == Decimal("0")
    assert rec.reason == "COORDINATOR_LEFT"
    assert any(line.role == "Team Lead" and line.eligible and line.amount == Decimal("250") for line in lines)
    assert any(line.role == "Manager" and line.eligible for line in lines)
    assert any(line.role == "CRM" and line.eligible for line in lines)


def test_nashik_manager_left_both_sides_continue_via_coordinator_master():
    db = _session()
    _seed_candidate(
        db,
        recruiter="Rec Active",
        team_lead="TL Active",
        manager="Mgr Left",
        crm="CRM Active",
        center_head=None,
        avp=None,
        margin=Decimal("8"),
    )
    _seed_coordinator(db, "Rec Active", CoordinatorStatus.ACTIVE, "1")
    _seed_coordinator(db, "TL Active", CoordinatorStatus.ACTIVE, "2")
    _seed_coordinator(db, "Mgr Left", CoordinatorStatus.LEFT, "3")
    _seed_coordinator(db, "CRM Active", CoordinatorStatus.ACTIVE, "4")
    cycle = _cycle(db)
    rows = [HoursMatchRow(uploaded_name="Aisha Mayes", uploaded_id="12345", hours=160)]
    lines, _, _, _ = run_cycle_calculation(db, cycle, rows, WINDOW)

    assert any(line.role == "Manager" and line.reason == "COORDINATOR_LEFT" and line.amount == 0 for line in lines)
    assert any(line.role == "Recruiter" and line.eligible and line.amount == Decimal("2000") for line in lines)
    assert any(line.role == "Team Lead" and line.eligible for line in lines)
    assert any(line.role == "CRM" and line.eligible for line in lines)


def test_nashik_nitin_three_roles_top_two_excludes_recruiter_in_cycle():
    db = _session()
    person = "Nitin Giri"
    _seed_candidate(
        db,
        recruiter=person,
        team_lead="Other TL",
        manager=person,
        crm=person,
        center_head=None,
        avp=None,
        margin=Decimal("8"),
    )
    _seed_coordinator(db, person, CoordinatorStatus.ACTIVE, "1")
    _seed_coordinator(db, "Other TL", CoordinatorStatus.ACTIVE, "2")
    cycle = _cycle(db)
    rows = [HoursMatchRow(uploaded_name="Aisha Mayes", uploaded_id="12345", hours=160)]
    lines, _, _, _ = run_cycle_calculation(db, cycle, rows, WINDOW)

    nitin = [line for line in lines if line.person == person and line.eligible and line.amount > 0]
    roles = {line.role for line in nitin}
    assert roles == {"Manager", "CRM"}
    assert not any(line.role == "Recruiter" and line.eligible and line.amount > 0 and line.person == person for line in lines)
