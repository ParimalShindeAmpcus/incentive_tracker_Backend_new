"""Regenerate Nashik left/hierarchy scenario Excel files with the working hours format."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

BACKEND_ROOT = Path(__file__).resolve().parents[1]
OUT = BACKEND_ROOT / "tests" / "test_data"
OUT.mkdir(parents=True, exist_ok=True)

CANDIDATE_HEADERS = [
    "Start ID",
    "Activity ID",
    "Candidate",
    "Email",
    "Contact",
    "Client",
    "End Client",
    "End Date",
    "Req ID",
    "Contract Type",
    "Subcontractor",
    "Subcontractor Email",
    "Subcontractor Contact",
    "Job Level",
    "Salary",
    "Pay Rate",
    "Taxes",
    "Benefits",
    "Referral Fee",
    "Gross Bill Rate",
    "MSP Fee",
    "Remote",
    "Work Location",
    "Candidate Location",
    "Work Authorization",
    "Resume Source",
    "Team Lead",
    "CRM",
    "Team Manager",
    "Senior Manager",
    "Associate Director",
    "Director",
    "Center Head",
    "Assistant VP",
    "Onboarding Coordinator",
    "Organization",
    "User Email",
    "Recruiter Location",
    "Job Title",
    "Start Date",
    "Margin",
    "Recruiter",
    "Status",
]

RECRUITER_STATUS_HEADERS = [
    "Coordinator Name",
    "Email",
    "Organization",
    "Role / Title",
    "Employment Status",
    "Exit Date",
    "Bank Name",
    "Account Number",
    "IFSC Code",
]


def candidate_row(**kwargs) -> list[object]:
    base = {h: "" for h in CANDIDATE_HEADERS}
    base.update(
        {
            "Client": "Acme Corp",
            "End Client": "Acme Corp",
            "Req ID": "REQ-NSK-LEFT",
            "Organization": "Ampcus Inc",
            "Resume Source": "Ampcus Inc",
            "Recruiter Location": "Nashik",
            "Remote": "Yes",
            "Work Location": "Remote",
            "Status": "Active",
            "Start Date": "2026-01-15",
            "Job Title": "Consultant",
            "Contract Type": "W2",
            "Margin": 8,
            "Pay Rate": 45,
            "Gross Bill Rate": 53,
        }
    )
    base.update(kwargs)
    if not base.get("Activity ID"):
        base["Activity ID"] = base.get("Start ID", "")
    return [base[h] for h in CANDIDATE_HEADERS]


def main() -> None:
    # Fresh Start IDs so upload creates new rows (does not collide with old W2-100x UAT data)
    candidates = [
        candidate_row(
            **{
                "Start ID": "NSK-LEFT-001",
                "Candidate": "Kavya Patil",
                "Email": "kavya.left.test@example.com",
                "Recruiter": "Ravi Left",
                "Team Lead": "Majid Khan",
            }
        ),
        candidate_row(
            **{
                "Start ID": "NSK-LEFT-002",
                "Candidate": "Manager Left Case",
                "Email": "mgr.left.test@example.com",
                "Recruiter": "Nitin Left",
                "Team Lead": "Majid Khan",
                "Team Manager": "Nitin Left",
                "CRM": "David",
                "Center Head": "ABC",
                "Assistant VP": "DEF",
            }
        ),
        candidate_row(
            **{
                "Start ID": "NSK-LEFT-003",
                "Candidate": "CRM Left Case",
                "Email": "crm.left.test@example.com",
                "Recruiter": "David Left",
                "Team Lead": "Majid Khan",
                "Team Manager": "Nitin Giri",
                "CRM": "David Left",
                "Center Head": "ABC",
                "Assistant VP": "DEF",
            }
        ),
        candidate_row(
            **{
                "Start ID": "NSK-LEFT-004",
                "Candidate": "Senior Manager Left Case",
                "Email": "sm.left.test@example.com",
                "Recruiter": "Senior Left",
                "Team Lead": "Majid Khan",
                "Team Manager": "Nitin Giri",
                "Senior Manager": "Senior Left",
                "CRM": "David",
                "Center Head": "ABC",
                "Assistant VP": "DEF",
            }
        ),
        candidate_row(
            **{
                "Start ID": "NSK-LEFT-005",
                "Candidate": "AVP Left Case",
                "Email": "avp.left.test@example.com",
                "Recruiter": "Avp Left",
                "Team Lead": "Majid Khan",
                "Team Manager": "Nitin Giri",
                "CRM": "David",
                "Center Head": "ABC",
                "Assistant VP": "Avp Left",
            }
        ),
        candidate_row(
            **{
                "Start ID": "NSK-LEFT-006",
                "Candidate": "Sameer Dual",
                "Email": "sameer.dual.test@example.com",
                "Recruiter": "Alex Dual",
                "Team Lead": "Alex Dual",
                "Team Manager": "Alex Dual",
                "CRM": "Alex Dual",
            }
        ),
    ]

    cand_path = OUT / "nashik_candidate_master_left_hierarchy_cases.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate_Master"
    ws.append(CANDIDATE_HEADERS)
    for row in candidates:
        ws.append(row)
    wb.save(cand_path)

    # Working hours format used by division-aware + reconciled exports:
    # Candidate Name, Candidate Start ID, Client Name, Hours Worked, Month
    # Month = August-2026 (not 2026-08)
    hours_headers = ["Candidate Name", "Candidate Start ID", "Client Name", "Hours Worked", "Month"]
    hours_rows = [
        ["Kavya Patil", "NSK-LEFT-001", "Acme Corp", 160, "August-2026"],
        ["Manager Left Case", "NSK-LEFT-002", "Acme Corp", 160, "August-2026"],
        ["CRM Left Case", "NSK-LEFT-003", "Acme Corp", 160, "August-2026"],
        ["Senior Manager Left Case", "NSK-LEFT-004", "Acme Corp", 160, "August-2026"],
        ["AVP Left Case", "NSK-LEFT-005", "Acme Corp", 160, "August-2026"],
        ["Sameer Dual", "NSK-LEFT-006", "Acme Corp", 160, "August-2026"],
    ]

    hours_path = OUT / "nashik_hours_august_left_hierarchy_cases.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Hours Template"
    ws2.append(hours_headers)
    for row in hours_rows:
        ws2.append(row)
    wb2.save(hours_path)

    csv_path = OUT / "nashik_hours_august_left_hierarchy_cases.csv"
    csv_path.write_text(
        ",".join(hours_headers)
        + "\n"
        + "\n".join(",".join(str(v) for v in row) for row in hours_rows)
        + "\n",
        encoding="utf-8",
    )

    status_path = OUT / "nashik_recruiter_status_left_hierarchy_cases.xlsx"
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "Recruiter_Status"
    ws3.append(RECRUITER_STATUS_HEADERS)
    for row in [
        ["Ravi Left", "ravi.left@gmail.com", "Ampcus Inc", "Recruiter", "LEFT", "2026-06-30", "HDFC", "1111111118", "HDFC0000001"],
        ["Nitin Left", "nitin.left@gmail.com", "Ampcus Inc", "Manager", "LEFT", "2026-06-30", "HDFC", "1111111113", "HDFC0000001"],
        ["David Left", "david.left@gmail.com", "Ampcus Inc", "CRM", "LEFT", "2026-06-30", "HDFC", "1111111114", "HDFC0000001"],
        ["Senior Left", "senior.left@gmail.com", "Ampcus Inc", "Senior Manager", "LEFT", "2026-06-30", "HDFC", "1111111117", "HDFC0000001"],
        ["Avp Left", "avp.left@gmail.com", "Ampcus Inc", "AVP", "LEFT", "2026-06-30", "HDFC", "1111111116", "HDFC0000001"],
        ["Alex Dual", "alex.dual@gmail.com", "Ampcus Inc", "Recruiter", "ACTIVE", "", "HDFC", "1111111120", "HDFC0000001"],
        ["Majid Khan", "majid.khan@gmail.com", "Ampcus Inc", "Team Lead", "ACTIVE", "", "HDFC", "1111111112", "HDFC0000001"],
        ["Nitin Giri", "nitin.giri@gmail.com", "Ampcus Inc", "Manager", "ACTIVE", "", "HDFC", "1111111121", "HDFC0000001"],
        ["David", "david.crm@gmail.com", "Ampcus Inc", "CRM", "ACTIVE", "", "HDFC", "1111111122", "HDFC0000001"],
        ["ABC", "abc.ch@gmail.com", "Ampcus Inc", "Center Head", "ACTIVE", "", "HDFC", "1111111123", "HDFC0000001"],
        ["DEF", "def.avp@gmail.com", "Ampcus Inc", "AVP", "ACTIVE", "", "HDFC", "1111111124", "HDFC0000001"],
    ]:
        ws3.append(row)
    wb3.save(status_path)

    print("Updated files:")
    print(f"- {cand_path}")
    print(f"- {hours_path}")
    print(f"- {csv_path}")
    print(f"- {status_path}")


if __name__ == "__main__":
    main()
