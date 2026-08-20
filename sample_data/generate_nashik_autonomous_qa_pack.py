"""
Generate unique Nashik autonomous QA pack (Candidate Master + Hours + Recruiter Status).

IDs are timestamped so they never collide with existing Candidate Master rows.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

BACKEND_ROOT = Path(__file__).resolve().parents[1]
OUT = BACKEND_ROOT / "tests" / "test_data"
OUT.mkdir(parents=True, exist_ok=True)

STAMP = datetime.now().strftime("%Y%m%d%H%M%S")
PREFIX = f"NASH-TEST-{STAMP}"

CANDIDATE_HEADERS = [
    "Start ID", "Activity ID", "Candidate", "Email", "Contact", "Client", "End Client", "End Date",
    "Req ID", "Contract Type", "Subcontractor", "Subcontractor Email", "Subcontractor Contact",
    "Job Level", "Salary", "Pay Rate", "Taxes", "Benefits", "Referral Fee", "Gross Bill Rate",
    "MSP Fee", "Remote", "Work Location", "Candidate Location", "Work Authorization", "Resume Source",
    "Team Lead", "CRM", "Team Manager", "Senior Manager", "Associate Director", "Director",
    "Center Head", "Assistant VP", "Onboarding Coordinator", "Organization", "User Email",
    "Recruiter Location", "Job Title", "Start Date", "Margin", "Recruiter", "Status",
]

STATUS_HEADERS = [
    "Coordinator Name", "Email", "Organization", "Role / Title", "Employment Status",
    "Exit Date", "Bank Name", "Account Number", "IFSC Code",
]


def emp(n: int, label: str) -> str:
    return f"NASH-EMP-{n:03d} {label} {STAMP}"


# Synthetic employees
E = {
    "rec_active": emp(1, "Recruiter Active"),
    "tl_active": emp(2, "Team Lead Active"),
    "mgr_active": emp(3, "Manager Active"),
    "crm_active": emp(4, "CRM Active"),
    "rec_left": emp(5, "Recruiter Left"),
    "tl_left": emp(6, "Team Lead Left"),
    "mgr_left": emp(7, "Manager Left"),
    "crm_left": emp(8, "CRM Left"),
    "ch_active": emp(9, "Center Head Active"),
    "nitin": emp(10, "Nitin Dual Roles"),
    "avp_active": emp(11, "AVP Active"),
    "all_roles": emp(12, "All Roles Person"),
}


def cand_row(sid: str, name: str, **hier) -> list[object]:
    base = {h: "" for h in CANDIDATE_HEADERS}
    base.update(
        {
            "Start ID": sid,
            "Activity ID": f"ACT-{sid}",
            "Candidate": name,
            "Email": f"{sid.lower().replace('-', '.')}@gmail.com",
            "Client": "Acme Corp",
            "End Client": "Acme Corp",
            "Req ID": f"REQ-{sid}",
            "Contract Type": "W2",
            "Organization": "Ampcus Inc",
            "Resume Source": "Ampcus Inc",
            "Recruiter Location": "Nashik",
            "Remote": "Yes",
            "Work Location": "Remote",
            "Status": "Active",
            "Start Date": "2026-01-15",
            "Job Title": "Consultant",
            "Margin": 8,
            "Pay Rate": 45,
            "Gross Bill Rate": 53,
            "Recruiter": hier.get("recruiter", ""),
            "Team Lead": hier.get("team_lead", ""),
            "Team Manager": hier.get("manager", ""),
            "Senior Manager": hier.get("senior_manager", ""),
            "CRM": hier.get("crm", ""),
            "Center Head": hier.get("center_head", ""),
            "Assistant VP": hier.get("avp", ""),
        }
    )
    return [base[h] for h in CANDIDATE_HEADERS]


def main() -> None:
    scenarios = [
        # A / G baseline all active
        {
            "id": f"{PREFIX}-001",
            "name": f"NASH Baseline All Active {STAMP}",
            "scenario": "A/G Normal hierarchy all active",
            "hier": {
                "recruiter": E["rec_active"],
                "team_lead": E["tl_active"],
                "manager": E["mgr_active"],
                "crm": E["crm_active"],
                "center_head": E["ch_active"],
                "avp": E["avp_active"],
            },
        },
        # B recruiter left
        {
            "id": f"{PREFIX}-002",
            "name": f"NASH Recruiter Left {STAMP}",
            "scenario": "B Recruiter left",
            "hier": {
                "recruiter": E["rec_left"],
                "team_lead": E["tl_active"],
                "manager": E["mgr_active"],
                "crm": E["crm_active"],
                "center_head": E["ch_active"],
            },
        },
        # C team lead left
        {
            "id": f"{PREFIX}-003",
            "name": f"NASH Team Lead Left {STAMP}",
            "scenario": "C Team Lead left",
            "hier": {
                "recruiter": E["rec_active"],
                "team_lead": E["tl_left"],
                "manager": E["mgr_active"],
                "crm": E["crm_active"],
            },
        },
        # D manager left
        {
            "id": f"{PREFIX}-004",
            "name": f"NASH Manager Left {STAMP}",
            "scenario": "D Manager left",
            "hier": {
                "recruiter": E["rec_active"],
                "team_lead": E["tl_active"],
                "manager": E["mgr_left"],
                "crm": E["crm_active"],
                "center_head": E["ch_active"],
            },
        },
        # E CRM left
        {
            "id": f"{PREFIX}-005",
            "name": f"NASH CRM Left {STAMP}",
            "scenario": "E CRM left",
            "hier": {
                "recruiter": E["rec_active"],
                "team_lead": E["tl_active"],
                "manager": E["mgr_active"],
                "crm": E["crm_left"],
                "center_head": E["ch_active"],
            },
        },
        # F multiple left
        {
            "id": f"{PREFIX}-006",
            "name": f"NASH Multi Left {STAMP}",
            "scenario": "F Recruiter+Manager left, TL+CRM active",
            "hier": {
                "recruiter": E["rec_left"],
                "team_lead": E["tl_active"],
                "manager": E["mgr_left"],
                "crm": E["crm_active"],
            },
        },
        # H same person Recruiter+CRM+Manager
        {
            "id": f"{PREFIX}-007",
            "name": f"NASH Dual Rec CRM Mgr {STAMP}",
            "scenario": "H Same employee Recruiter+CRM+Manager",
            "hier": {
                "recruiter": E["nitin"],
                "team_lead": E["tl_active"],
                "manager": E["nitin"],
                "crm": E["nitin"],
            },
        },
        # I Recruiter+TL+Manager
        {
            "id": f"{PREFIX}-008",
            "name": f"NASH Dual Rec TL Mgr {STAMP}",
            "scenario": "I Same employee Recruiter+TeamLead+Manager",
            "hier": {
                "recruiter": E["nitin"],
                "team_lead": E["nitin"],
                "manager": E["nitin"],
                "crm": E["crm_active"],
            },
        },
        # J all hierarchy roles same person
        {
            "id": f"{PREFIX}-009",
            "name": f"NASH All Roles Same Person {STAMP}",
            "scenario": "J Same employee in all hierarchy roles",
            "hier": {
                "recruiter": E["all_roles"],
                "team_lead": E["all_roles"],
                "manager": E["all_roles"],
                "crm": E["all_roles"],
                "center_head": E["all_roles"],
                "avp": E["all_roles"],
            },
        },
        # N missing hierarchy
        {
            "id": f"{PREFIX}-010",
            "name": f"NASH Missing Hierarchy {STAMP}",
            "scenario": "N Missing Manager/CRM fields",
            "hier": {
                "recruiter": E["rec_active"],
                "team_lead": E["tl_active"],
            },
        },
        # O inactive employee with hours (recruiter left)
        {
            "id": f"{PREFIX}-011",
            "name": f"NASH Left Rec With Hours {STAMP}",
            "scenario": "O Left recruiter with 160h",
            "hier": {
                "recruiter": E["rec_left"],
                "team_lead": E["tl_active"],
            },
        },
        # P active with zero hours handled via hours file row of 0 — keep candidate for master
        {
            "id": f"{PREFIX}-012",
            "name": f"NASH Zero Hours Active {STAMP}",
            "scenario": "P Active hierarchy with 0 hours",
            "hier": {
                "recruiter": E["rec_active"],
                "team_lead": E["tl_active"],
                "manager": E["mgr_active"],
            },
            "hours": 0,
        },
    ]

    # Candidate Master
    cand_path = OUT / f"nashik_candidate_master_test_data_{STAMP}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate_Master"
    ws.append(CANDIDATE_HEADERS)
    for s in scenarios:
        ws.append(cand_row(s["id"], s["name"], **s["hier"]))
    guide = wb.create_sheet("Scenarios")
    guide.append(["Start ID", "Candidate", "Scenario Code"])
    for s in scenarios:
        guide.append([s["id"], s["name"], s["scenario"]])
    emp_sheet = wb.create_sheet("Employees")
    emp_sheet.append(["Employee Key", "Display Name", "Status Intent"])
    for key, name in E.items():
        status = "LEFT" if "Left" in name or "left" in key else "ACTIVE"
        emp_sheet.append([key, name, status])
    wb.save(cand_path)

    # Hours reconciled
    hours_path = OUT / f"nashik_hours_reconciled_test_data_{STAMP}.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Hours Template"
    ws2.append(["Candidate Name", "Candidate Start ID", "Client Name", "Hours Worked", "Month"])
    for s in scenarios:
        hours = s.get("hours", 160)
        ws2.append([s["name"], s["id"], "Acme Corp", hours, "August-2026"])
    audit = wb2.create_sheet("Audit Detail")
    audit.append(["Candidate Name", "Candidate Start ID", "Client Name", "Hours Worked", "Month", "Scenario"])
    for s in scenarios:
        audit.append([s["name"], s["id"], "Acme Corp", s.get("hours", 160), "August-2026", s["scenario"]])
    wb2.save(hours_path)

    csv_path = OUT / f"nashik_hours_reconciled_test_data_{STAMP}.csv"
    lines = ["Candidate Name,Candidate Start ID,Client Name,Hours Worked,Month"]
    for s in scenarios:
        lines.append(f"{s['name']},{s['id']},Acme Corp,{s.get('hours', 160)},August-2026")
    csv_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    # Recruiter status
    status_path = OUT / f"nashik_recruiter_status_test_data_{STAMP}.xlsx"
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "Recruiter_Status"
    ws3.append(STATUS_HEADERS)
    status_rows = [
        (E["rec_active"], "Recruiter", "ACTIVE"),
        (E["tl_active"], "Team Lead", "ACTIVE"),
        (E["mgr_active"], "Manager", "ACTIVE"),
        (E["crm_active"], "CRM", "ACTIVE"),
        (E["rec_left"], "Recruiter", "LEFT"),
        (E["tl_left"], "Team Lead", "LEFT"),
        (E["mgr_left"], "Manager", "LEFT"),
        (E["crm_left"], "CRM", "LEFT"),
        (E["ch_active"], "Center Head", "ACTIVE"),
        (E["nitin"], "Manager", "ACTIVE"),
        (E["avp_active"], "AVP", "ACTIVE"),
        (E["all_roles"], "Recruiter", "ACTIVE"),
    ]
    for i, (name, role, status) in enumerate(status_rows, start=1):
        email = f"nash.emp.{STAMP}.{i}@gmail.com"
        exit_date = "2026-06-30" if status == "LEFT" else ""
        ws3.append([name, email, "Ampcus Inc", role, status, exit_date, "HDFC", f"{STAMP}{i:02d}"[-12:], "HDFC0000001"])
    wb3.save(status_path)

    # Expected results sheet (repo-authoritative + business-gap notes)
    expected_path = OUT / f"nashik_expected_results_{STAMP}.xlsx"
    wb4 = Workbook()
    ws4 = wb4.active
    ws4.title = "Expected_Repo_Behavior"
    ws4.append(
        [
            "Scenario",
            "Candidate ID",
            "Repo Expected Lines (backend calculate_nashik_placement)",
            "Business Requirement (prompt)",
            "Gap?",
        ]
    )
    expected_rows = [
        [
            "A/G Baseline",
            f"{PREFIX}-001",
            "Recruiter 2000; TL 250; top2 hierarchy by priority among Mgr/CRM/CH/AVP => AVP 2300 + Center Head 1500 (Mgr/CRM dropped by max-2). Total eligible 2000+250+2300+1500=6050",
            "All roles processed",
            "YES — repo caps hierarchy at 2 roles/person; AVP+CH win over Mgr+CRM",
        ],
        [
            "B Recruiter left",
            f"{PREFIX}-002",
            "BACKEND: Recruiter LEFT still paid 2000 (no LEFT check). Hierarchy continues.",
            "Recruiter excluded; hierarchy continues",
            "YES — backend Nashik ignores LEFT status",
        ],
        [
            "C TL left",
            f"{PREFIX}-003",
            "BACKEND: TL LEFT still paid 250. Recruiter+Mgr+CRM continue.",
            "TL excluded; others continue",
            "YES — no LEFT filter on Nashik engine",
        ],
        [
            "D Manager left",
            f"{PREFIX}-004",
            "BACKEND: Manager LEFT still paid if selected in top2. Others continue.",
            "Manager excluded; others continue",
            "YES",
        ],
        [
            "E CRM left",
            f"{PREFIX}-005",
            "BACKEND: CRM LEFT still paid if selected.",
            "CRM excluded; others continue",
            "YES",
        ],
        [
            "F Multi left",
            f"{PREFIX}-006",
            "BACKEND: Left people still paid. Candidate not discarded.",
            "Inactive excluded; active remain",
            "YES on exclusion; PASS that candidate continues",
        ],
        [
            "H Rec+CRM+Mgr same person",
            f"{PREFIX}-007",
            "Recruiter 2000 for Nitin + hierarchy top2 of (Mgr,CRM,TL)= Manager 1500 + CRM 1000. TL active separate. Total Nitin roles=3 lines",
            "Only Manager+CRM; NO Recruiter",
            "YES — repo still pays Recruiter separately; max-2 only on hierarchy",
        ],
        [
            "I Rec+TL+Mgr same",
            f"{PREFIX}-008",
            "Recruiter 2000 + hierarchy top2 of (TL,Mgr)= Manager 1500 + TL 250. CRM active separate.",
            "Top 2 roles only total",
            "PARTIAL — hierarchy capped; Recruiter still extra",
        ],
        [
            "J All roles same person",
            f"{PREFIX}-009",
            "Recruiter 2000 + hierarchy top2 AVP 2300 + Center Head 1500 (Mgr/CRM/TL dropped)",
            "Exactly 2 roles total",
            "YES — 3 lines (Recruiter + 2 hierarchy)",
        ],
        [
            "N Missing hierarchy",
            f"{PREFIX}-010",
            "Recruiter 2000 + TL 250 only. No crash. No fake Mgr/CRM.",
            "No crash; remaining processed",
            "NO gap",
        ],
        [
            "O Left with hours",
            f"{PREFIX}-011",
            "BACKEND: Left recruiter still eligible 2000 if hours 160",
            "Hours alone must not override LEFT ineligibility",
            "YES",
        ],
        [
            "P Zero hours",
            f"{PREFIX}-012",
            "Recruiter pro-rata 0; TL pro-rata 0; Manager one-time ineligible (<160)",
            "No manufactured incentive",
            "NO gap",
        ],
    ]
    for row in expected_rows:
        ws4.append(row)
    notes = wb4.create_sheet("Source_Of_Truth")
    notes.append(["Topic", "Location"])
    notes.append(["Amounts / slabs", "app/services/incentives/nashik_rules.py"])
    notes.append(["Calculator", "app/services/incentives/nashik_calculator.py"])
    notes.append(["Cycle engine dispatch", "app/services/cycles/cycle_engine.py (is_nashik_division branch)"])
    notes.append(["LEFT handling (Client/InHouse/Sambhaji only)", "engines/ampcus_client.py, ampcus_inhouse.py, sambhaji_nagar.py"])
    notes.append(["Frontend LEFT recruiter only", "src/services/incentiveCalculator.ts + recruiterStatusService.ts"])
    notes.append(["Approve / export", "cycle_service.approve_cycle + export approved excel"])
    wb4.save(expected_path)

    readme = OUT / f"nashik_cycle_test_readme_{STAMP}.txt"
    readme.write_text(
        "\n".join(
            [
                f"Nashik autonomous QA pack {STAMP}",
                "",
                "Upload order:",
                f"1. {cand_path.name}",
                f"2. {status_path.name}",
                "3. Create Nashik cycle August 2026",
                f"4. Upload hours: {hours_path.name}",
                "5. Calculate (and optionally Approve)",
                "",
                "Also run unit tests:",
                "  pytest tests/unit/test_nashik_autonomous_scenarios.py -q",
                "",
                f"Expected results workbook: {expected_path.name}",
            ]
        ),
        encoding="utf-8",
    )

    print(cand_path)
    print(hours_path)
    print(csv_path)
    print(status_path)
    print(expected_path)
    print(readme)
    print(f"PREFIX={PREFIX}")
    print(f"STAMP={STAMP}")


if __name__ == "__main__":
    main()
