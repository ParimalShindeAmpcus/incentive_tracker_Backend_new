"""Generate unique Nashik LEFT/hierarchy scenario files that will not merge with existing master."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

BACKEND_ROOT = Path(__file__).resolve().parents[1]
OUT = BACKEND_ROOT / "tests" / "test_data"
OUT.mkdir(parents=True, exist_ok=True)

STAMP = datetime.now().strftime("%Y%m%d%H%M")
PREFIX = f"NSK-U{STAMP}"

CANDIDATE_HEADERS = [
    "Start ID",
    "Activity ID",
    "Candidate",
    "Email",
    "Contact",
    "Client",
    "End Client",
    "End Date",
    "Req ID",
    "Contract Type",
    "Subcontractor",
    "Subcontractor Email",
    "Subcontractor Contact",
    "Job Level",
    "Salary",
    "Pay Rate",
    "Taxes",
    "Benefits",
    "Referral Fee",
    "Gross Bill Rate",
    "MSP Fee",
    "Remote",
    "Work Location",
    "Candidate Location",
    "Work Authorization",
    "Resume Source",
    "Team Lead",
    "CRM",
    "Team Manager",
    "Senior Manager",
    "Associate Director",
    "Director",
    "Center Head",
    "Assistant VP",
    "Onboarding Coordinator",
    "Organization",
    "User Email",
    "Recruiter Location",
    "Job Title",
    "Start Date",
    "Margin",
    "Recruiter",
    "Status",
]

RECRUITER_STATUS_HEADERS = [
    "Coordinator Name",
    "Email",
    "Organization",
    "Role / Title",
    "Employment Status",
    "Exit Date",
    "Bank Name",
    "Account Number",
    "IFSC Code",
]


def candidate_row(**kwargs) -> list[object]:
    base = {h: "" for h in CANDIDATE_HEADERS}
    base.update(
        {
            "Client": "Acme Corp",
            "End Client": "Acme Corp",
            "Req ID": f"REQ-{PREFIX}",
            "Organization": "Ampcus Inc",
            "Resume Source": "Ampcus Inc",
            "Recruiter Location": "Nashik",
            "Remote": "Yes",
            "Work Location": "Remote",
            "Status": "Active",
            "Start Date": "2026-01-15",
            "Job Title": "Consultant",
            "Contract Type": "W2",
            "Margin": 8,
            "Pay Rate": 45,
            "Gross Bill Rate": 53,
        }
    )
    base.update(kwargs)
    if not base.get("Activity ID"):
        base["Activity ID"] = base.get("Start ID", "")
    return [base[h] for h in CANDIDATE_HEADERS]


def main() -> None:
    # Unique people for this run only (timestamped IDs + unique names)
    scenarios = [
        {
            "sid": f"{PREFIX}-01",
            "name": f"Unique Left Recruiter Cand {STAMP}",
            "email": f"unique.left.rec.{STAMP}@gmail.com",
            "recruiter": f"Unique Ravi Left {STAMP}",
            "team_lead": f"Unique Majid TL {STAMP}",
            "manager": "",
            "senior_manager": "",
            "crm": "",
            "center_head": "",
            "avp": "",
            "case": "S1 Left recruiter - hierarchy still paid",
        },
        {
            "sid": f"{PREFIX}-02",
            "name": f"Unique Manager Left Cand {STAMP}",
            "email": f"unique.mgr.left.{STAMP}@gmail.com",
            "recruiter": f"Unique Active Rec {STAMP}",
            "team_lead": f"Unique Majid TL {STAMP}",
            "manager": f"Unique Manager Left {STAMP}",
            "senior_manager": "",
            "crm": f"Unique David CRM {STAMP}",
            "center_head": f"Unique Center Head {STAMP}",
            "avp": f"Unique AVP Active {STAMP}",
            "case": "S2 Manager left - others continue",
        },
        {
            "sid": f"{PREFIX}-03",
            "name": f"Unique CRM Left Cand {STAMP}",
            "email": f"unique.crm.left.{STAMP}@gmail.com",
            "recruiter": f"Unique Active Rec {STAMP}",
            "team_lead": f"Unique Majid TL {STAMP}",
            "manager": f"Unique Manager Active {STAMP}",
            "senior_manager": "",
            "crm": f"Unique CRM Left {STAMP}",
            "center_head": f"Unique Center Head {STAMP}",
            "avp": f"Unique AVP Active {STAMP}",
            "case": "S2b CRM left - others continue",
        },
        {
            "sid": f"{PREFIX}-04",
            "name": f"Unique Senior Mgr Left Cand {STAMP}",
            "email": f"unique.sm.left.{STAMP}@gmail.com",
            "recruiter": f"Unique Active Rec {STAMP}",
            "team_lead": f"Unique Majid TL {STAMP}",
            "manager": f"Unique Manager Active {STAMP}",
            "senior_manager": f"Unique Senior Left {STAMP}",
            "crm": f"Unique David CRM {STAMP}",
            "center_head": f"Unique Center Head {STAMP}",
            "avp": f"Unique AVP Active {STAMP}",
            "case": "S2c Senior Manager left - others continue",
        },
        {
            "sid": f"{PREFIX}-05",
            "name": f"Unique AVP Left Cand {STAMP}",
            "email": f"unique.avp.left.{STAMP}@gmail.com",
            "recruiter": f"Unique Active Rec {STAMP}",
            "team_lead": f"Unique Majid TL {STAMP}",
            "manager": f"Unique Manager Active {STAMP}",
            "senior_manager": "",
            "crm": f"Unique David CRM {STAMP}",
            "center_head": f"Unique Center Head {STAMP}",
            "avp": f"Unique AVP Left {STAMP}",
            "case": "S2d AVP left - others continue",
        },
        {
            "sid": f"{PREFIX}-06",
            "name": f"Unique Dual Role Cand {STAMP}",
            "email": f"unique.dual.{STAMP}@gmail.com",
            "recruiter": f"Unique Dual Person {STAMP}",
            "team_lead": f"Unique Dual Person {STAMP}",
            "manager": f"Unique Dual Person {STAMP}",
            "senior_manager": "",
            "crm": f"Unique Dual Person {STAMP}",
            "center_head": "",
            "avp": "",
            "case": "S3 Same person in 3+ roles - max 2 only",
        },
    ]

    cand_rows = []
    for s in scenarios:
        cand_rows.append(
            candidate_row(
                **{
                    "Start ID": s["sid"],
                    "Activity ID": f"ACT-{s['sid']}",
                    "Candidate": s["name"],
                    "Email": s["email"],
                    "Recruiter": s["recruiter"],
                    "Team Lead": s["team_lead"],
                    "Team Manager": s["manager"],
                    "Senior Manager": s["senior_manager"],
                    "CRM": s["crm"],
                    "Center Head": s["center_head"],
                    "Assistant VP": s["avp"],
                }
            )
        )

    cand_path = OUT / f"nashik_unique_candidate_master_{STAMP}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate_Master"
    ws.append(CANDIDATE_HEADERS)
    for row in cand_rows:
        ws.append(row)
    # Guide sheet
    guide = wb.create_sheet("Test_Cases")
    guide.append(["Start ID", "Candidate", "Scenario"])
    for s in scenarios:
        guide.append([s["sid"], s["name"], s["case"]])
    wb.save(cand_path)

    hours_headers = ["Candidate Name", "Candidate Start ID", "Client Name", "Hours Worked", "Month"]
    hours_rows = [
        [s["name"], s["sid"], "Acme Corp", 160, "August-2026"] for s in scenarios
    ]
    hours_path = OUT / f"nashik_unique_hours_reconciled_{STAMP}.xlsx"
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Hours Template"
    ws2.append(hours_headers)
    for row in hours_rows:
        ws2.append(row)
    audit = wb2.create_sheet("Audit Detail")
    audit.append(hours_headers + ["Match Status", "Notes"])
    for row, s in zip(hours_rows, scenarios):
        audit.append(list(row) + ["expected_match", s["case"]])
    wb2.save(hours_path)

    csv_path = OUT / f"nashik_unique_hours_reconciled_{STAMP}.csv"
    csv_path.write_text(
        ",".join(hours_headers)
        + "\n"
        + "\n".join(",".join(str(v) for v in row) for row in hours_rows)
        + "\n",
        encoding="utf-8",
    )

    # Unique coordinator / recruiter status names for LEFT cases
    left_people = [
        (f"Unique Ravi Left {STAMP}", f"unique.ravi.left.{STAMP}@gmail.com", "Recruiter", "LEFT"),
        (f"Unique Manager Left {STAMP}", f"unique.mgr.left.person.{STAMP}@gmail.com", "Manager", "LEFT"),
        (f"Unique CRM Left {STAMP}", f"unique.crm.left.person.{STAMP}@gmail.com", "CRM", "LEFT"),
        (f"Unique Senior Left {STAMP}", f"unique.sm.left.person.{STAMP}@gmail.com", "Senior Manager", "LEFT"),
        (f"Unique AVP Left {STAMP}", f"unique.avp.left.person.{STAMP}@gmail.com", "AVP", "LEFT"),
        (f"Unique Dual Person {STAMP}", f"unique.dual.person.{STAMP}@gmail.com", "Recruiter", "ACTIVE"),
        (f"Unique Active Rec {STAMP}", f"unique.active.rec.{STAMP}@gmail.com", "Recruiter", "ACTIVE"),
        (f"Unique Majid TL {STAMP}", f"unique.majid.tl.{STAMP}@gmail.com", "Team Lead", "ACTIVE"),
        (f"Unique Manager Active {STAMP}", f"unique.mgr.active.{STAMP}@gmail.com", "Manager", "ACTIVE"),
        (f"Unique David CRM {STAMP}", f"unique.david.crm.{STAMP}@gmail.com", "CRM", "ACTIVE"),
        (f"Unique Center Head {STAMP}", f"unique.ch.{STAMP}@gmail.com", "Center Head", "ACTIVE"),
        (f"Unique AVP Active {STAMP}", f"unique.avp.active.{STAMP}@gmail.com", "AVP", "ACTIVE"),
    ]

    status_path = OUT / f"nashik_unique_recruiter_status_{STAMP}.xlsx"
    wb3 = Workbook()
    ws3 = wb3.active
    ws3.title = "Recruiter_Status"
    ws3.append(RECRUITER_STATUS_HEADERS)
    for i, (name, email, role, status) in enumerate(left_people, start=1):
        exit_date = "2026-06-30" if status == "LEFT" else ""
        ws3.append(
            [
                name,
                email,
                "Ampcus Inc",
                role,
                status,
                exit_date,
                "HDFC",
                f"9{STAMP}{i:02d}"[-12:],
                "HDFC0000001",
            ]
        )
    wb3.save(status_path)

    readme = OUT / f"nashik_unique_TEST_README_{STAMP}.txt"
    readme.write_text(
        "\n".join(
            [
                f"Nashik unique LEFT / hierarchy test pack — {STAMP}",
                "",
                "Upload order:",
                f"1) Candidate Master: {cand_path.name}",
                f"2) Recruiter / Coordinator status: {status_path.name}",
                "3) Create Nashik cycle for August 2026",
                f"4) Cycle Step 1 hours: {hours_path.name}",
                "5) Calculate and review results",
                "",
                "Start IDs:",
                *[f"  {s['sid']}  {s['name']}  ->  {s['case']}" for s in scenarios],
                "",
                "Expected:",
                "S1: Left recruiter gets 0; Team Lead still paid",
                "S2/S2b/S2c/S2d: Left higher authority withheld; other hierarchy still paid; names still on cycle lines",
                "S3: Dual person paid for at most 2 highest roles only",
            ]
        ),
        encoding="utf-8",
    )

    print("Generated unique pack:")
    print(cand_path)
    print(hours_path)
    print(csv_path)
    print(status_path)
    print(readme)
    print(f"PREFIX={PREFIX}")


if __name__ == "__main__":
    main()
