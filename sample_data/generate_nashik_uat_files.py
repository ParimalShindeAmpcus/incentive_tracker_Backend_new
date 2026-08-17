"""Generate Nashik UAT upload files: candidates, coordinators, hours, scenarios."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
NOTE_FILL = PatternFill("solid", fgColor="FFF4CC")
PASS_FILL = PatternFill("solid", fgColor="EAF6EA")
WARN_FILL = PatternFill("solid", fgColor="FFF4CC")
FAIL_FILL = PatternFill("solid", fgColor="FDECEC")
THIN = Border(
    left=Side(style="thin", color="D0D5DD"),
    right=Side(style="thin", color="D0D5DD"),
    top=Side(style="thin", color="D0D5DD"),
    bottom=Side(style="thin", color="D0D5DD"),
)

CANDIDATE_HEADERS = [
    "Start ID",
    "Activity ID",
    "Candidate",
    "Email",
    "Contact",
    "Client",
    "End Client",
    "End Date",
    "Req ID",
    "Contract Type",
    "Subcontractor",
    "Subcontractor Email",
    "Subcontractor Contact",
    "Job Level",
    "Salary",
    "Pay Rate",
    "Taxes",
    "Benefits",
    "Referral Fee",
    "Gross Bill Rate",
    "MSP Fee",
    "Remote",
    "Work Location",
    "Candidate Location",
    "Work Authorization",
    "Resume Source",
    "Team Lead",
    "CRM",
    "Team Manager",
    "Senior Manager",
    "Associate Director",
    "Director",
    "Center Head",
    "Assistant VP",
    "Onboarding Coordinator",
    "Organization",
    "User Email",
    "Recruiter Location",
    "Job Title",
    "Start Date",
    "Margin",
    "Recruiter",
    "Status",
]


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = THIN
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


def autosize(ws, max_w=32):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max((len(str(c.value or "")) for c in col), default=12) + 3, max_w)
        ws.column_dimensions[letter].width = max(width, 14)


def write_rows(ws, headers, rows, fills=None):
    ws.append(headers)
    for i, row in enumerate(rows):
        values = [row.get(h, "") for h in headers] if isinstance(row, dict) else list(row)
        ws.append(values)
        fill = (fills or [None])[i] if fills else None
        for col in range(1, len(values) + 1):
            cell = ws.cell(i + 2, col)
            cell.border = THIN
            if fill:
                cell.fill = fill
    style_header(ws, len(headers))
    autosize(ws)


def candidate_row(**kwargs):
    base = {h: "" for h in CANDIDATE_HEADERS}
    base.update(
        {
            "Email": "",
            "Contact": "",
            "Client": "Acme Corp",
            "End Client": "Acme Corp",
            "Req ID": "REQ-1",
            "Organization": "Ampcus Inc",
            "Resume Source": "Ampcus Inc",
            "Recruiter Location": "Nashik",
            "Remote": "Yes",
            "Work Location": "Remote",
            "Status": "Active",
            "Start Date": "2026-01-15",
            "Job Title": "Consultant",
        }
    )
    base.update(kwargs)
    if not base.get("Activity ID"):
        base["Activity ID"] = base.get("Start ID", "")
    return base


candidates = [
    candidate_row(
        **{
            "Start ID": "W2-1001",
            "Candidate": "Priya Sharma",
            "Contract Type": "W2",
            "Margin": 8,
            "Pay Rate": 45,
            "Gross Bill Rate": 53,
            "Recruiter": "Amit Ohol",
            "Team Lead": "Majid Khan",
            "Email": "priya.sharma@example.com",
        }
    ),
    candidate_row(
        **{
            "Start ID": "12345",
            "Candidate": "Aisha Mayes",
            "Contract Type": "C2C",
            "Margin": 12,
            "Pay Rate": 70,
            "Gross Bill Rate": 82,
            "Start Date": "2026-01-01",
            "Recruiter": "Amit Ohol",
            "Team Lead": "Majid Khan",
            "Team Manager": "Nitin Giri",
            "CRM": "David",
            "Center Head": "ABC",
            "Assistant VP": "DEF",
            "Subcontractor": "Mayes LLC",
            "Client": "Globex",
            "End Client": "Globex",
            "Email": "aisha.mayes@example.com",
        }
    ),
    candidate_row(
        **{
            "Start ID": "W2-1002",
            "Candidate": "Rohan Desai",
            "Contract Type": "W2",
            "Margin": 8,
            "Recruiter": "Amit Ohol",
            "Team Lead": "Majid Khan",
            "Email": "rohan.desai@example.com",
        }
    ),
    candidate_row(
        **{
            "Start ID": "W2-1003",
            "Candidate": "Kavya Patil",
            "Contract Type": "W2",
            "Margin": 8,
            "Recruiter": "Ravi Left",
            "Team Lead": "Majid Khan",
            "Email": "kavya.patil@example.com",
        }
    ),
    candidate_row(
        **{
            "Start ID": "W2-1004",
            "Candidate": "Sameer Dual",
            "Contract Type": "W2",
            "Margin": 8,
            "Recruiter": "Alex Dual",
            "Team Lead": "Alex Dual",
            "Team Manager": "Alex Dual",
            "CRM": "Alex Dual",
            "Email": "sameer.dual@example.com",
        }
    ),
    candidate_row(
        **{
            "Start ID": "W2-1005",
            "Candidate": "Neha Joshi",
            "Contract Type": "W2",
            "Margin": 8,
            "Recruiter": "Amit Ohol",
            "Email": "neha.joshi@example.com",
        }
    ),
    candidate_row(
        **{
            "Start ID": "C2C-1006",
            "Candidate": "Low Margin C2C",
            "Contract Type": "C2C",
            "Margin": 0.99,
            "Recruiter": "Amit Ohol",
            "Team Lead": "Majid Khan",
            "Email": "low.margin@example.com",
        }
    ),
    candidate_row(
        **{
            "Start ID": "W2-1007",
            "Candidate": "Early End Case",
            "Contract Type": "W2",
            "Margin": 8,
            "End Date": "2026-08-10",
            "Recruiter": "Amit Ohol",
            "Team Lead": "Majid Khan",
            "Team Manager": "Nitin Giri",
            "CRM": "David",
            "Email": "early.end@example.com",
        }
    ),
    candidate_row(
        **{
            "Start ID": "W2-1008",
            "Candidate": "Cumulative Lead",
            "Contract Type": "W2",
            "Margin": 8,
            "Recruiter": "Amit Ohol",
            "Team Lead": "Majid Khan",
            "Team Manager": "Nitin Giri",
            "CRM": "David",
            "Email": "cumulative.lead@example.com",
        }
    ),
    candidate_row(
        **{
            "Start ID": "W2-1009",
            "Candidate": "Mismatch Person",
            "Contract Type": "W2",
            "Margin": 8,
            "Recruiter": "Amit Ohol",
            "Team Lead": "Majid Khan",
            "Email": "mismatch.person@example.com",
        }
    ),
    candidate_row(
        **{
            "Start ID": "W2-1010",
            "Candidate": "Vikram Kulkarni",
            "Contract Type": "W2",
            "Margin": 8,
            "Recruiter": "Amit Ohol",
            "Team Lead": "Majid Khan",
            "Email": "vikram.kulkarni@example.com",
        }
    ),
    candidate_row(
        **{
            "Start ID": "W2-1011",
            "Candidate": "Notice Case",
            "Contract Type": "W2",
            "Margin": 8,
            "Recruiter": "Sonia Notice",
            "Team Lead": "Majid Khan",
            "Email": "notice.case@example.com",
        }
    ),
]


def save_candidates():
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate_Master"
    write_rows(ws, CANDIDATE_HEADERS, candidates)
    path = OUT / "nashik_test_candidates.xlsx"
    wb.save(path)
    return path


COORD_HEADERS = [
    "Coordinator Name",
    "Email",
    "Organization",
    "Role",
    "Employment Status",
    "Exit Date",
    "Bank Name",
    "Account Number",
    "IFSC Code",
]

coordinators = [
    ["Amit Ohol", "amit.ohol@example.com", "Ampcus Inc", "Recruiter", "ACTIVE", "", "HDFC", "1111111111", "HDFC0000001"],
    ["Majid Khan", "majid.khan@example.com", "Ampcus Inc", "Team Lead", "ACTIVE", "", "HDFC", "1111111112", "HDFC0000001"],
    ["Nitin Giri", "nitin.giri@example.com", "Ampcus Inc", "Manager", "ACTIVE", "", "HDFC", "1111111113", "HDFC0000001"],
    ["David", "david.crm@example.com", "Ampcus Inc", "CRM", "ACTIVE", "", "HDFC", "1111111114", "HDFC0000001"],
    ["ABC", "abc.ch@example.com", "Ampcus Inc", "Center Head", "ACTIVE", "", "HDFC", "1111111115", "HDFC0000001"],
    ["DEF", "def.avp@example.com", "Ampcus Inc", "AVP", "ACTIVE", "", "HDFC", "1111111116", "HDFC0000001"],
    ["Alex Dual", "alex.dual@example.com", "Ampcus Inc", "Recruiter", "ACTIVE", "", "HDFC", "1111111117", "HDFC0000001"],
    ["Ravi Left", "ravi.left@example.com", "Ampcus Inc", "Recruiter", "LEFT", "2026-06-30", "HDFC", "1111111118", "HDFC0000001"],
    ["Sonia Notice", "sonia.notice@example.com", "Ampcus Inc", "Recruiter", "NOTICE", "", "HDFC", "1111111119", "HDFC0000001"],
]


def save_coordinators():
    wb = Workbook()
    ws = wb.active
    ws.title = "Coordinators"
    write_rows(
        ws,
        COORD_HEADERS,
        coordinators,
        fills=[PASS_FILL] * 7 + [FAIL_FILL, WARN_FILL],
    )
    path = OUT / "nashik_test_coordinators.xlsx"
    wb.save(path)
    return path


HOURS_HEADERS = ["Candidate ID", "Candidate Name", "Client Name", "Hours Worked", "Month"]

hours_august = [
    ["W2-1001", "Priya Sharma", "Acme Corp", 160, "2026-08"],
    ["12345", "Aisha Mayes", "Globex", 160, "2026-08"],
    ["W2-1002", "Rohan Desai", "Acme Corp", 80, "2026-08"],
    ["W2-1003", "Kavya Patil", "Acme Corp", 160, "2026-08"],
    ["W2-1004", "Sameer Dual", "Acme Corp", 160, "2026-08"],
    ["W2-1005", "Neha Joshi", "Acme Corp", 160, "2026-08"],
    ["C2C-1006", "Low Margin C2C", "Acme Corp", 160, "2026-08"],
    ["W2-1007", "Early End Case", "Acme Corp", 80, "2026-08"],
    ["W2-1008", "Cumulative Lead", "Acme Corp", 80, "2026-08"],
    ["WRONG-999", "Mismatch Person", "Acme Corp", 160, "2026-08"],
    ["W2-1010", "Vikram Kulkarnee", "Acme Corp", 160, "2026-08"],
    ["W2-1011", "Notice Case", "Acme Corp", 160, "2026-08"],
]

hours_july = [
    ["W2-1008", "Cumulative Lead", "Acme Corp", 80, "2026-07"],
]


def save_hours(name, rows, title):
    wb = Workbook()
    ws = wb.active
    ws.title = title
    write_rows(ws, HOURS_HEADERS, rows)
    path = OUT / name
    wb.save(path)
    return path


def save_scenarios():
    wb = Workbook()
    ws = wb.active
    ws.title = "How_To_Test"
    steps = [
        ["Step", "Action", "File"],
        ["1", "Upload Candidate Master / Starts", "nashik_test_candidates.xlsx"],
        ["2", "Upload Coordinators (bulk)", "nashik_test_coordinators.xlsx"],
        ["3", "Optional — first create a July Nashik cycle, upload July hours, Calculate (do not need to approve)", "nashik_test_hours_july.xlsx"],
        ["4", "Create Nashik cycle for August 2026", "Division = nashik, month = 2026-08"],
        ["5", "Upload August hours on the cycle and click Calculate", "nashik_test_hours_august.xlsx"],
        ["6", "Compare results to sheet Expected_Results", "this workbook"],
    ]
    write_rows(ws, steps[0], steps[1:])
    ws.merge_cells("A9:C12")
    ws["A9"] = (
        "Use a saved (database) Nashik cycle so calculation runs on the backend. "
        "W2 and C2C use the same slabs. Hours file has no contract/margin — those come from Candidate Master. "
        "July hours file is only required for scenario S9 (cumulative 160h)."
    )
    ws["A9"].fill = NOTE_FILL
    ws["A9"].alignment = Alignment(wrap_text=True, vertical="top")

    exp = wb.create_sheet("Expected_Results")
    headers = [
        "Scenario",
        "Candidate ID",
        "Candidate Name",
        "What you are testing",
        "Hours file row",
        "Expected match",
        "Expected eligible payouts (INR)",
        "Should NOT pay",
    ]
    expected = [
        [
            "S1 Happy path W2 160h",
            "W2-1001",
            "Priya Sharma",
            "W2 slab $8 → recruiter 2000 + TL 250",
            "Name and ID both correct, 160h",
            "Name + ID",
            "Amit Ohol Recruiter 2000 Recurring; Majid Khan Team Lead 250 Recurring. Total 2250",
            "No Manager/CRM/AVP (blank on Master)",
        ],
        [
            "S2 Happy path C2C 160h",
            "12345",
            "Aisha Mayes",
            "Same Nashik slabs as W2; margin $12; full hierarchy",
            "Name and ID both correct, 160h",
            "Name + ID",
            "Recruiter 3500; TL 250; Manager 1500; CRM 1000; Center Head 1500; AVP 2300. Total 10050",
            "",
        ],
        [
            "S3 Pro-rata 80h",
            "W2-1002",
            "Rohan Desai",
            "Recruiter and TL pro-rata hours/160",
            "80 hours",
            "Name + ID",
            "Recruiter 1000; Team Lead 125. Total 1125",
            "Leadership one-time (below 160 cumulative)",
        ],
        [
            "S4 Left recruiter",
            "W2-1003",
            "Kavya Patil",
            "Coordinator Ravi Left = LEFT — block his payout",
            "160 hours",
            "Name + ID",
            "Majid Khan Team Lead 250 only",
            "Ravi Left Recruiter 0",
        ],
        [
            "S5 Max 2 roles",
            "W2-1004",
            "Sameer Dual",
            "Same person Recruiter + TL + Manager + CRM — keep top 2 amounts",
            "160 hours",
            "Name + ID",
            "Alex Dual Recruiter 2000 and Alex Dual Manager 1500 only. Total 3500",
            "Team Lead 250 and CRM 1000 dropped (3rd/4th role)",
        ],
        [
            "S6 Missing hierarchy",
            "W2-1005",
            "Neha Joshi",
            "Do not invent Team Lead / leadership",
            "160 hours",
            "Name + ID",
            "Amit Ohol Recruiter 2000 only",
            "No TL/Manager/CRM rows",
        ],
        [
            "S7 Low margin special",
            "C2C-1006",
            "Low Margin C2C",
            "Margin 0.99 → recruiter one-time 2000, not slab",
            "160 hours",
            "Name + ID",
            "Recruiter SPECIAL 2000; TL 250",
            "Normal 160h slab",
        ],
        [
            "S8 Project end before 160",
            "W2-1007",
            "Early End Case",
            "End Date 2026-08-10 and only 80h this cycle / cumulative < 160",
            "80 hours",
            "Name + ID",
            "Recruiter SPECIAL 2000 only",
            "TL 0; Manager/CRM 0",
        ],
        [
            "S9 Cumulative 160h",
            "W2-1008",
            "Cumulative Lead",
            "July 80h calculated first, then August 80h",
            "August file 80h after July cycle exists",
            "Name + ID",
            "August recruiter pro-rata 1000; TL 125; Manager 1500 one-time; CRM 1000 one-time",
            "If July was never calculated: leadership stays 0",
        ],
        [
            "S10 Name/ID mismatch",
            "W2-1009",
            "Mismatch Person",
            "Hours name matches Master, hours ID is WRONG-999",
            "ID WRONG-999, name Mismatch Person",
            "Mismatch — do not calculate",
            "0",
            "Any incentive",
        ],
        [
            "S11 ID fallback",
            "W2-1010",
            "Vikram Kulkarni",
            "Hours name is misspelled, ID is correct",
            "Name Vikram Kulkarnee, ID W2-1010",
            "ID fallback warning + still calculate",
            "Recruiter 2000; TL 250",
            "",
        ],
        [
            "S12 Notice recruiter",
            "W2-1011",
            "Notice Case",
            "Sonia Notice = NOTICE — block her payout",
            "160 hours",
            "Name + ID",
            "Majid Khan Team Lead 250 only",
            "Sonia Notice Recruiter 0",
        ],
    ]
    write_rows(exp, headers, expected)

    notes = wb.create_sheet("Notes")
    notes.append(["Topic", "Rule"])
    notes.append(["W2 vs C2C", "Same Nashik margin slabs. Difference is Master margin and who is filled in hierarchy."])
    notes.append(["Hours file", "Only Candidate ID, Candidate Name, Client Name, Hours Worked, Month. Do not put margin/contract here."])
    notes.append(["Matching", "Name first, then ID. Name match + wrong ID = no pay. Wrong name + unique ID = warning + pay."])
    notes.append(["Cumulative", "Leadership 160h uses earlier cycle hours + this month. Recruiter/TL still use this month only."])
    notes.append(["Max 2 roles", "Per person per placement, keep the two highest amounts, including Recruiter."])
    notes.append(["Left/Notice", "Uses Coordinator master employment status. Names must match hierarchy names."])
    write_rows(notes, ["Topic", "Rule"], [notes.iter_rows(min_row=2, values_only=True)])  # will duplicate; rewrite cleanly

    path = OUT / "nashik_test_scenarios.xlsx"
    wb.save(path)
    return path


def save_scenarios_clean():
    wb = Workbook()
    ws = wb.active
    ws.title = "How_To_Test"
    how = [
        ["Step", "Action", "File"],
        ["1", "Upload Candidate Master / Starts", "nashik_test_candidates.xlsx"],
        ["2", "Upload Coordinators (bulk)", "nashik_test_coordinators.xlsx"],
        ["3", "Optional: create July Nashik cycle, upload July hours, Calculate", "nashik_test_hours_july.xlsx"],
        ["4", "Create Nashik cycle for August 2026 (division nashik)", "—"],
        ["5", "Upload August hours and Calculate", "nashik_test_hours_august.xlsx"],
        ["6", "Compare results to Expected_Results", "nashik_test_scenarios.xlsx"],
    ]
    write_rows(ws, how[0], how[1:])
    ws.merge_cells("A9:C12")
    ws["A9"] = (
        "Use a saved database cycle so calculation runs on the backend. "
        "W2 and C2C use the same slabs. Hours file does not contain contract type or margin. "
        "July hours are required only for scenario S9 (cumulative 160 hours)."
    )
    ws["A9"].fill = NOTE_FILL
    ws["A9"].alignment = Alignment(wrap_text=True, vertical="top")

    exp = wb.create_sheet("Expected_Results")
    headers = [
        "Scenario",
        "Candidate ID",
        "Candidate Name",
        "What you are testing",
        "Hours file row",
        "Expected match",
        "Expected eligible payouts (INR)",
        "Should NOT pay",
    ]
    expected = [
        ["S1 Happy path W2 160h", "W2-1001", "Priya Sharma", "W2 slab $8", "160h correct name+ID", "Name + ID", "Recruiter Amit Ohol 2000; TL Majid Khan 250. Total 2250", "No Manager/CRM/AVP"],
        ["S2 Happy path C2C 160h", "12345", "Aisha Mayes", "C2C same slabs, margin $12, full hierarchy", "160h correct name+ID", "Name + ID", "Recruiter 3500; TL 250; Manager 1500; CRM 1000; CH 1500; AVP 2300. Total 10050", ""],
        ["S3 Pro-rata 80h", "W2-1002", "Rohan Desai", "Pro-rata hours/160", "80h", "Name + ID", "Recruiter 1000; TL 125. Total 1125", "Leadership one-time"],
        ["S4 Left recruiter", "W2-1003", "Kavya Patil", "Ravi Left = LEFT", "160h", "Name + ID", "TL Majid Khan 250", "Ravi Left Recruiter 0"],
        ["S5 Max 2 roles", "W2-1004", "Sameer Dual", "Same person Recruiter+TL+Manager+CRM", "160h", "Name + ID", "Alex Dual Recruiter 2000 and Manager 1500 only. Total 3500", "TL 250 and CRM 1000 dropped"],
        ["S6 Missing hierarchy", "W2-1005", "Neha Joshi", "Do not invent people", "160h", "Name + ID", "Recruiter 2000 only", "No TL/leadership rows"],
        ["S7 Low margin", "C2C-1006", "Low Margin C2C", "Margin 0.99 special 2000", "160h", "Name + ID", "Recruiter SPECIAL 2000; TL 250", "Normal slab"],
        ["S8 Project end <160", "W2-1007", "Early End Case", "End Date 2026-08-10 and 80h", "80h", "Name + ID", "Recruiter SPECIAL 2000 only", "TL and leadership 0"],
        ["S9 Cumulative 160h", "W2-1008", "Cumulative Lead", "July 80h cycle then August 80h", "August 80h after July calculated", "Name + ID", "Aug Recruiter 1000; TL 125; Manager 1500; CRM 1000", "If July never calculated, leadership = 0"],
        ["S10 Name/ID mismatch", "W2-1009", "Mismatch Person", "Name matches, ID WRONG-999", "ID WRONG-999", "Mismatch — no calculate", "0", "Any incentive"],
        ["S11 ID fallback", "W2-1010", "Vikram Kulkarni", "Wrong name, correct ID", "Name Vikram Kulkarnee", "ID fallback warning + calculate", "Recruiter 2000; TL 250", ""],
        ["S12 Notice recruiter", "W2-1011", "Notice Case", "Sonia Notice = NOTICE", "160h", "Name + ID", "TL Majid Khan 250", "Sonia Notice Recruiter 0"],
    ]
    write_rows(exp, headers, expected)

    notes = wb.create_sheet("Notes")
    write_rows(
        notes,
        ["Topic", "Rule"],
        [
            ["W2 vs C2C", "Same Nashik margin slabs. Difference is Master margin and filled hierarchy."],
            ["Hours file", "Only Candidate ID, Candidate Name, Client Name, Hours Worked, Month."],
            ["Matching", "Name first, then ID. Name+wrong ID = no pay. Wrong name+unique ID = warning + pay."],
            ["Cumulative", "Leadership uses earlier cycle hours + this month. Recruiter/TL use this month only."],
            ["Max 2 roles", "Per person per placement keep the two highest amounts, including Recruiter."],
            ["Left/Notice", "Coordinator master employment status. Names must match Candidate Master hierarchy."],
        ],
    )
    path = OUT / "nashik_test_scenarios.xlsx"
    wb.save(path)
    return path


def main():
    paths = [
        save_candidates(),
        save_coordinators(),
        save_hours("nashik_test_hours_august.xlsx", hours_august, "Hours"),
        save_hours("nashik_test_hours_july.xlsx", hours_july, "Hours"),
        save_scenarios_clean(),
    ]
    for path in paths:
        print(path)


if __name__ == "__main__":
    main()
