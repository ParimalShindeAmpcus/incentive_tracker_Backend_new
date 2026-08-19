"""Parse the filled Candidate Hours template (Excel/CSV)."""

from __future__ import annotations

import csv
import io
from decimal import Decimal, InvalidOperation
from typing import List

from openpyxl import load_workbook

from app.services.cycles.hours_name_matcher import HoursMatchRow


def _header_key(value: object) -> str:
    return " ".join(str(value or "").strip().lower().split())


def _cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _hours(value: object) -> Decimal:
    raw = _cell(value).replace(",", "")
    if not raw:
        return Decimal("0")
    try:
        return Decimal(raw)
    except InvalidOperation:
        return Decimal("0")


COLUMN_ALIASES = {
    "candidate name": "name",
    "candidate": "name",
    "consultant": "name",
    "name": "name",
    "candidate start id": "id",
    "candidate id": "id",
    "start id": "id",
    "id": "id",
    "client name": "client",
    "client": "client",
    "hours worked": "hours",
    "hours": "hours",
    "month": "month",
    "period": "month",
    "billing month": "month",
    "incentive month": "month",
}


def parse_hours_template(
    content: bytes,
    filename: str = "hours.xlsx",
    *,
    require_hours: bool = True,
) -> List[HoursMatchRow]:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _parse_csv(content, require_hours=require_hours)
    return _parse_xlsx(content, require_hours=require_hours)


def _map_headers(headers: List[str], *, require_hours: bool = True) -> dict:
    mapping = {}
    for idx, header in enumerate(headers):
        alias = COLUMN_ALIASES.get(_header_key(header))
        if alias and alias not in mapping:
            mapping[alias] = idx
    required = ("name", "id", "client", "month")
    if require_hours:
        required = ("name", "id", "hours")
    labels = {
        "name": "Candidate Name",
        "id": "Candidate ID",
        "hours": "Hours Worked",
        "client": "Client Name",
        "month": "Month",
    }
    missing = [labels[label] for label in required if label not in mapping]
    if missing:
        kind = "Hours file" if require_hours else "Placement file"
        raise ValueError(f"{kind} is missing required columns: {', '.join(missing)}")
    return mapping


def _rows_from_values(
    headers: List[str],
    data_rows: List[List[object]],
    *,
    require_hours: bool = True,
) -> List[HoursMatchRow]:
    mapping = _map_headers(headers, require_hours=require_hours)
    out: List[HoursMatchRow] = []
    for offset, values in enumerate(data_rows, start=2):
        def take(key: str) -> str:
            idx = mapping.get(key)
            if idx is None or idx >= len(values):
                return ""
            return _cell(values[idx])

        name = take("name")
        ident = take("id")
        if not name and not ident:
            continue
        hours_raw = None
        if "hours" in mapping:
            hours_idx = mapping["hours"]
            hours_raw = values[hours_idx] if hours_idx < len(values) else None
        out.append(
            HoursMatchRow(
                uploaded_name=name,
                uploaded_id=ident,
                client=take("client"),
                hours=float(_hours(hours_raw)),
                month=take("month"),
                source_row=offset,
            )
        )
    if not out:
        kind = "hours file" if require_hours else "placement file"
        raise ValueError(f"No data rows found in the {kind}")
    return out


def _parse_xlsx(content: bytes, *, require_hours: bool = True) -> List[HoursMatchRow]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Hours Excel file is empty")
    headers = [_cell(v) for v in rows[0]]
    return _rows_from_values(headers, [list(r) for r in rows[1:]], require_hours=require_hours)


def _parse_csv(content: bytes, *, require_hours: bool = True) -> List[HoursMatchRow]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("Hours CSV file is empty")
    return _rows_from_values(rows[0], rows[1:], require_hours=require_hours)
