import csv, io
from openpyxl import load_workbook
from typing import Optional
from fastapi import HTTPException, status
from pydantic import ValidationError
from sqlalchemy.orm import Session
from app.models.coordinators.schemas import CoordinatorInput, CoordinatorStatusUpdate, CoordinatorUpdate
from app.repositories.coordinators import coordinator_repository as repo
from app.repositories.entities.coordinator import CoordinatorRecord, CoordinatorStatus

def norm(value: str): return " ".join(value.strip().lower().split())
def apply_status(record, status_value, exit_date):
    if exit_date:
        # Check if exit_date is a string and parse it, or if it is already a date object
        from datetime import datetime
        d_val = exit_date
        if isinstance(d_val, str):
            try:
                d_val = datetime.strptime(d_val, "%Y-%m-%d").date()
            except ValueError:
                try:
                    d_val = datetime.fromisoformat(d_val).date()
                except ValueError:
                    pass
        if isinstance(d_val, date) and d_val <= date.today():
            record.employment_status = CoordinatorStatus.LEFT
        else:
            record.employment_status = status_value
    else:
        record.employment_status = status_value
    record.exit_date = exit_date
    record.incentive_eligible = record.employment_status == CoordinatorStatus.ACTIVE

def list_coordinators(db: Session, page: int, page_size: int, search: Optional[str], employment_status: Optional[CoordinatorStatus]):
    return {"items": repo.list_records(db, (page-1)*page_size, page_size, search, employment_status), "total": repo.count(db, search, employment_status), "page": page, "page_size": page_size}
def get_coordinator(db, record_id):
    record = repo.get(db, record_id)
    if not record: raise HTTPException(status_code=404, detail="Coordinator not found")
    return record
def summary(db):
    counts = repo.counts(db); return {"total_coordinators": sum(counts.values()), "active_coordinators": counts["ACTIVE"], "left_coordinators": counts["LEFT"], "notice_period_coordinators": counts["NOTICE"], "incentive_eligible_coordinators": counts["ACTIVE"]}
def create(db: Session, payload: CoordinatorInput):
    email = str(payload.email).lower()
    if repo.by_email(db, email): raise HTTPException(status_code=409, detail="A coordinator with this email already exists")
    record = CoordinatorRecord(full_name=payload.full_name.strip(), normalized_name=norm(payload.full_name), email=email, organization=payload.organization.strip(), role_title=payload.role_title.strip(), start_date=payload.start_date, bank_name=payload.bank_name, account_number=payload.account_number, ifsc_code=payload.ifsc_code.upper() if payload.ifsc_code else None)
    apply_status(record, payload.employment_status, payload.exit_date); db.add(record); db.commit(); db.refresh(record); return record
def update(db: Session, record_id: int, payload: CoordinatorUpdate):
    record = get_coordinator(db, record_id); data = payload.model_dump(exclude_unset=True)
    if "email" in data and data["email"]:
        existing = repo.by_email(db, str(data["email"]).lower())
        if existing and existing.id != record.id: raise HTTPException(status_code=409, detail="A coordinator with this email already exists")
        record.email = str(data.pop("email")).lower()
    if "full_name" in data and data["full_name"]: record.full_name=data["full_name"].strip(); record.normalized_name=norm(record.full_name)
    for field in ("organization", "role_title", "start_date", "bank_name", "account_number", "ifsc_code"):
        if field in data: setattr(record, field, data[field].strip().upper() if field == "ifsc_code" and data[field] else data[field])
    apply_status(record, data.get("employment_status", record.employment_status), data.get("exit_date", record.exit_date)); db.commit(); db.refresh(record); return record
def update_status(db, record_id, payload: CoordinatorStatusUpdate): return update(db, record_id, CoordinatorUpdate(employment_status=payload.employment_status, exit_date=payload.exit_date))
def delete_left(db, record_id):
    record = get_coordinator(db, record_id)
    if record.employment_status != CoordinatorStatus.LEFT: raise HTTPException(status_code=422, detail="Only coordinators marked Left can be deleted")
    record.is_deleted = True; db.commit()
def _rows(content: bytes, filename: str):
    if filename.lower().endswith(".xlsx"):
        sheet = load_workbook(io.BytesIO(content), data_only=True).active
        values = list(sheet.iter_rows(values_only=True))
        if not values:
            return []
        headers = [str(value or "").strip() for value in values[0]]
        return [dict(zip(headers, ["" if value is None else str(value) for value in row])) for row in values[1:]]
    return list(csv.DictReader(io.StringIO(content.decode("utf-8-sig"))))
def bulk_upload(db, content: bytes, filename: str):
    issues=[]; created=0
    for index,row in enumerate(_rows(content, filename), 2):
        try:
            payload=CoordinatorInput(full_name=row.get("Coordinator Name") or row.get("Full Name") or "", email=row.get("Email") or "", organization=row.get("Organization") or "", role_title=row.get("Role") or row.get("Role / Title") or "", employment_status=row.get("Employment Status") or "ACTIVE", exit_date=row.get("Exit Date") or None, bank_name=row.get("Bank Name") or None, account_number=row.get("Account Number") or None, ifsc_code=row.get("IFSC Code") or None)
            create(db,payload); created+=1
        except (ValidationError, HTTPException) as exc: issues.append({"source_row":index,"identifier":row.get("Email") or row.get("Coordinator Name") or f"Row {index}","reason":str(getattr(exc,"detail",exc))})
    return {"created_count":created,"issues":issues}
def bulk_mark_left(db, content: bytes, filename: str):
    issues=[]; changed=0
    for index,row in enumerate(_rows(content, filename), 2):
        email=(row.get("Email") or "").strip().lower()
        if not email: issues.append({"source_row":index,"identifier":f"Row {index}","reason":"Email address is blank"}); continue
        record=repo.by_email(db,email)
        if not record: issues.append({"source_row":index,"identifier":email,"reason":"Email was not found"}); continue
        if record.employment_status != CoordinatorStatus.LEFT: apply_status(record,CoordinatorStatus.LEFT,None); changed+=1
    db.commit(); return {"marked_left_count":changed,"issues":issues}
